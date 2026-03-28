from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from itertools import combinations
from pathlib import Path
from typing import Callable, Iterable, Sequence
from zipfile import ZipFile

import numpy as np
from openpyxl import load_workbook

from official_cwl import OfficialMetrics, build_official_metrics, load_or_refresh_ssq_cache, parse_int


RED_MAX = 33
BLUE_MAX = 16
TICKET_COST = 2
BUDGETS = [10, 50, 100, 200, 500, 1000]

# ── 高回报约束 ──────────────────────────────────────────────────────────────
# 蓝球复式最多覆盖此数量的蓝球；超过即视为"宽蓝覆盖"，予以拒绝
BLUE_FUSHI_MAX_BLUES = 8
# 整体五等奖及以上命中率下限（低于此值则拒绝该投注方案）
MIN_VALUE_HIT_RATE_OVERALL = 0.030
# 近期五等奖及以上命中率下限
MIN_VALUE_HIT_RATE_RECENT = 0.020
# 整体四等奖及以上命中率下限（防止方案永远只命中六等奖）
MIN_STRONG_HIT_RATE_OVERALL = 0.004
# ───────────────────────────────────────────────────────────────────────────

FIXED_PAYOUTS = {
    "三等奖": 3000,
    "四等奖": 200,
    "五等奖": 10,
    "六等奖": 5,
}
FEATURE_NAMES = (
    "micro",
    "short",
    "mid",
    "long",
    "xl",
    "wide",
    "decay",
    "gap",
    "trend",
    "momentum",
    "rebound",
    "consistency",
    "similarity",
    # ── 混沌原理 & 概率论新增特征 ──────────────────────────────────────────
    "entropy",     # Shannon熵偏差：近期分布偏离均匀分布越大，说明该号码进入可预测态
    "markov",      # 一阶Markov转移概率：上期出现的号码对本期号码的条件概率之和
    "recurrence",  # 递归分析：在1~12期滞后窗口中出现的次数（识别周期性节律）
)

ARM_CONFIGS = {
    "balanced": {
        "micro": 0.10,
        "short": 0.28,
        "mid": 0.18,
        "long": 0.14,
        "xl": 0.08,
        "wide": 0.10,
        "decay": 0.22,
        "gap": -0.08,
        "trend": 0.14,
        "momentum": 0.05,
        "consistency": 0.05,
        "similarity": 0.24,  # 旧:0.18 — 相似期信号加强
    },
    "hot": {
        "micro": 0.24,
        "short": 0.46,
        "mid": 0.20,
        "long": 0.08,
        "xl": 0.00,
        "decay": 0.26,
        "gap": -0.20,
        "trend": 0.20,
        "momentum": 0.12,
        "similarity": 0.16,  # 旧:0.10
    },
    "cold_rebound": {
        "short": -0.10,      # 旧:-0.08 更强抑制近期热号
        "mid": 0.06,          # 旧:0.08
        "long": 0.12,         # 旧:0.15
        "xl": 0.15,           # 旧:0.18
        "wide": 0.10,         # 旧:0.12
        "decay": 0.04,        # 旧:0.05
        "gap": 0.55,          # 旧:0.50 ← 关键：更重视缺席期数（2026032: 11/31长期缺席后回弹）
        "trend": 0.04,        # 旧:0.05
        "rebound": 0.38,      # 旧:0.30 ← 关键：更重视回弹信号
        "consistency": 0.03,  # 旧:0.04
        "similarity": 0.16,   # 旧:0.14
    },
    "trend": {
        "micro": 0.18,
        "short": 0.12,
        "mid": -0.08,
        "long": 0.08,
        "xl": 0.02,
        "decay": 0.28,
        "gap": 0.02,
        "trend": 0.52,
        "momentum": 0.25,
        "similarity": 0.18,  # 旧:0.12
    },
    "steady": {
        "short": 0.06,
        "mid": 0.26,
        "long": 0.24,
        "xl": 0.10,
        "wide": 0.16,
        "decay": 0.12,
        "gap": 0.00,
        "trend": 0.04,
        "consistency": 0.16,
        "similarity": 0.28,  # 旧:0.20
    },
    "mean_revert": {
        "micro": -0.08,
        "short": -0.15,
        "mid": 0.12,
        "long": 0.14,
        "xl": 0.12,
        "wide": 0.10,
        "decay": 0.04,
        "gap": 0.50,
        "trend": -0.04,
        "rebound": 0.28,
        "consistency": 0.06,
        "similarity": 0.12,  # 旧:0.06
    },
    # ── 混沌原理臂：Shannon熵驱动 ─────────────────────────────────────────
    # 偏好近期分布偏离均匀态（低熵）的号码，捕捉"混沌吸引子附近"的局部可预测区域
    "chaos_entropy": {
        "entropy":     0.38,  # 熵偏差：低熵号码进入可预测态
        "markov":      0.22,  # Markov条件概率：从上期状态转移
        "recurrence":  0.16,  # 周期性递归信号
        "decay":       0.10,
        "consistency": 0.08,
        "similarity":  0.12,
    },
    # ── 概率论臂：Markov链 + 递归周期 ────────────────────────────────────
    # 依据一阶Markov转移矩阵和周期性递归，偏好条件概率最高的号码
    "markov_chain": {
        "markov":      0.40,  # 一阶Markov转移概率（主权重）
        "recurrence":  0.24,  # 滞后递归（周期律）
        "entropy":     0.10,  # 熵辅助
        "gap":         0.10,  # 间隔辅助
        "rebound":     0.08,
        "similarity":  0.08,
    },
}


def numbers_to_mask(numbers: Sequence[int]) -> int:
    mask = 0
    for number in numbers:
        mask |= 1 << (number - 1)
    return mask


@dataclass(frozen=True)
class Draw:
    issue: int
    draw_date: str
    reds: tuple[int, ...]
    blue: int
    official: OfficialMetrics | None = None
    red_mask: int = field(init=False, repr=False)

    def __post_init__(self) -> None:
        object.__setattr__(self, "red_mask", numbers_to_mask(self.reds))


@dataclass(frozen=True)
class Ticket:
    reds: tuple[int, ...]
    blue: int
    score: float
    red_mask: int = field(init=False, repr=False)

    def __post_init__(self) -> None:
        object.__setattr__(self, "red_mask", numbers_to_mask(self.reds))


@dataclass
class Scheme:
    family: str
    description: str
    cost: int
    tickets: list[Ticket]
    proxy_score: float
    metadata: dict[str, object] = field(default_factory=dict)


@dataclass
class SchemeResult:
    total_return: int
    profit: int
    prize_counts: Counter
    winning_ticket_count: int
    value_hit_tickets: int
    strong_hit_tickets: int
    max_red_hits: int
    blue_hit_tickets: int
    average_match_score: float
    utility: float
    ticket_returns: list[int] = field(default_factory=list)


@dataclass(frozen=True)
class ValidationRecord:
    cost: int
    total_return: int
    winning_ticket_count: int
    value_hit_tickets: int
    strong_hit_tickets: int
    blue_hit_tickets: int
    max_red_hits: int
    average_match_score: float
    ticket_count: int

    @property
    def any_prize_hit(self) -> bool:
        return self.winning_ticket_count > 0 or self.total_return > 0

    @property
    def value_hit(self) -> bool:
        return self.value_hit_tickets > 0

    @property
    def strong_hit(self) -> bool:
        return self.strong_hit_tickets > 0

    @property
    def ticket_hit_ratio(self) -> float:
        if self.ticket_count <= 0:
            return 0.0
        return self.winning_ticket_count / self.ticket_count

    @property
    def hit_score(self) -> float:
        issue_hit = 1.0 if self.any_prize_hit else 0.0
        value_hit = 1.0 if self.value_hit else 0.0
        strong_hit = 1.0 if self.strong_hit else 0.0
        coverage_component = min(self.ticket_hit_ratio * 5.0, 1.0)
        match_component = min(self.average_match_score / 6.5, 1.0)
        return (
            0.32 * issue_hit
            + 0.24 * value_hit
            + 0.16 * strong_hit
            + 0.18 * coverage_component
            + 0.10 * match_component
        )


@dataclass
class HistorySnapshot:
    issue: int
    context_vector: np.ndarray
    reds: tuple[int, ...]
    blue: int


@dataclass
class RewardHitSnapshot:
    issue: int
    budget: int
    family: str
    profile_key: str
    context_vector: np.ndarray
    reward_ratio: float
    profit_ratio: float


@dataclass
class BucketStat:
    count: int = 0
    payout_sum: float = 0.0

    @property
    def mean(self) -> float:
        return self.payout_sum / self.count if self.count else 0.0

    def update(self, payout: float) -> None:
        self.count += 1
        self.payout_sum += payout


@dataclass
class FamilyLiveState:
    count: int = 0
    ema_profit: float = 0.0
    ema_return: float = 0.0
    ema_hit: float = 0.0
    ema_value_hit: float = 0.0
    loss_streak: int = 0
    hitless_streak: int = 0
    cooldown: int = 0


@dataclass
class OnlineBinaryModel:
    feature_names: tuple[str, ...]
    learning_rate: float
    l2: float
    positive_weight: float
    weights: np.ndarray = field(init=False)
    bias: float = 0.0

    def __post_init__(self) -> None:
        self.weights = np.zeros(len(self.feature_names), dtype=float)

    def _matrix(self, feature_maps: dict[str, dict[int, float]]) -> tuple[list[int], np.ndarray]:
        numbers = sorted(next(iter(feature_maps.values())).keys())
        matrix = np.array(
            [[feature_maps[feature_name][number] for feature_name in self.feature_names] for number in numbers],
            dtype=float,
        )
        return numbers, matrix

    def predict_scores(self, feature_maps: dict[str, dict[int, float]]) -> dict[int, float]:
        numbers, matrix = self._matrix(feature_maps)
        logits = np.clip(matrix @ self.weights + self.bias, -18.0, 18.0)
        probabilities = 1.0 / (1.0 + np.exp(-logits))
        return normalize_scores({number: float(prob) for number, prob in zip(numbers, probabilities)})

    def update(self, feature_maps: dict[str, dict[int, float]], positive_numbers: set[int]) -> None:
        numbers, matrix = self._matrix(feature_maps)
        labels = np.array([1.0 if number in positive_numbers else 0.0 for number in numbers], dtype=float)
        logits = np.clip(matrix @ self.weights + self.bias, -18.0, 18.0)
        probabilities = 1.0 / (1.0 + np.exp(-logits))
        sample_weights = np.where(labels > 0.0, self.positive_weight, 1.0)
        errors = (probabilities - labels) * sample_weights
        grad_weights = (matrix.T @ errors) / len(numbers) + self.l2 * self.weights
        grad_bias = float(errors.mean())
        self.weights -= self.learning_rate * grad_weights
        self.bias -= self.learning_rate * grad_bias


@dataclass
class TicketReturnCalibrator:
    bucket_count: int = 24
    prior_strength: float = 18.0
    family_strength: float = 10.0
    global_stats: dict[int, BucketStat] = field(default_factory=lambda: defaultdict(BucketStat))
    family_stats: dict[str, dict[int, BucketStat]] = field(
        default_factory=lambda: defaultdict(lambda: defaultdict(BucketStat))
    )
    overall_stat: BucketStat = field(default_factory=BucketStat)
    family_overall_stats: dict[str, BucketStat] = field(default_factory=lambda: defaultdict(BucketStat))

    def bucket(self, score: float) -> int:
        clamped = min(max(score, 0.0), 0.999999)
        return int(clamped * self.bucket_count)

    def score_prior(self, score: float) -> float:
        return 0.35 + 0.45 * score + 0.55 * (score ** 2)

    def predict_ticket_return(self, family: str, ticket: Ticket) -> float:
        bucket = self.bucket(ticket.score)
        global_bucket = self.global_stats[bucket]
        family_bucket = self.family_stats[family][bucket]
        overall_mean = self.overall_stat.mean if self.overall_stat.count else self.score_prior(ticket.score)
        family_mean = (
            self.family_overall_stats[family].mean if self.family_overall_stats[family].count else overall_mean
        )
        numerator = self.prior_strength * self.score_prior(ticket.score)
        denominator = self.prior_strength
        numerator += max(2.0, math.sqrt(global_bucket.count + 1.0)) * (global_bucket.mean or overall_mean)
        denominator += max(2.0, math.sqrt(global_bucket.count + 1.0))
        numerator += max(1.0, math.sqrt(family_bucket.count + 1.0)) * (family_bucket.mean or family_mean)
        denominator += max(1.0, math.sqrt(family_bucket.count + 1.0))
        numerator += 3.0 * overall_mean + 2.0 * family_mean
        denominator += 5.0
        return numerator / denominator

    def predict_scheme_return(self, family: str, scheme: Scheme) -> float:
        return sum(self.predict_ticket_return(family, ticket) for ticket in scheme.tickets)

    def update_ticket(self, family: str, ticket: Ticket, payout: int) -> None:
        bucket = self.bucket(ticket.score)
        self.global_stats[bucket].update(payout)
        self.family_stats[family][bucket].update(payout)
        self.overall_stat.update(payout)
        self.family_overall_stats[family].update(payout)

    def update_scheme(self, family: str, scheme: Scheme, ticket_returns: Sequence[int]) -> None:
        for ticket, payout in zip(scheme.tickets, ticket_returns):
            self.update_ticket(family, ticket, payout)


@dataclass
class RoiStat:
    count: int = 0
    roi_sum: float = 0.0

    @property
    def mean(self) -> float:
        return self.roi_sum / self.count if self.count else 0.0

    def update(self, roi: float) -> None:
        self.count += 1
        self.roi_sum += roi


@dataclass
class SchemeGateCalibrator:
    bucket_count: int = 16
    prior_roi: float = -0.45  # 彩票理论期望回报约 -50%，使用更现实的先验避免早期过于乐观
    global_stats: dict[int, dict[int, RoiStat]] = field(default_factory=lambda: defaultdict(lambda: defaultdict(RoiStat)))
    family_stats: dict[int, dict[str, dict[int, RoiStat]]] = field(
        default_factory=lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(RoiStat)))
    )
    budget_overall_stats: dict[int, RoiStat] = field(default_factory=lambda: defaultdict(RoiStat))
    family_overall_stats: dict[int, dict[str, RoiStat]] = field(
        default_factory=lambda: defaultdict(lambda: defaultdict(RoiStat))
    )

    def bucket(self, confidence: float) -> int:
        clamped = min(max(confidence, 0.0), 0.999999)
        return int(clamped * self.bucket_count)

    def predict_roi(self, budget: int, family: str, confidence: float) -> float:
        bucket = self.bucket(confidence)
        global_bucket = self.global_stats[budget][bucket]
        family_bucket = self.family_stats[budget][family][bucket]
        budget_overall = self.budget_overall_stats[budget]
        family_overall = self.family_overall_stats[budget][family]
        budget_mean = budget_overall.mean if budget_overall.count else self.prior_roi
        family_mean = family_overall.mean if family_overall.count else budget_mean
        numerator = 4.0 * self.prior_roi
        denominator = 4.0
        numerator += max(1.5, math.sqrt(global_bucket.count + 1.0)) * (global_bucket.mean if global_bucket.count else budget_mean)
        denominator += max(1.5, math.sqrt(global_bucket.count + 1.0))
        numerator += max(2.5, 1.35 * math.sqrt(family_bucket.count + 1.0)) * (
            family_bucket.mean if family_bucket.count else family_mean
        )
        denominator += max(2.5, 1.35 * math.sqrt(family_bucket.count + 1.0))
        numerator += 1.0 * budget_mean + 3.0 * family_mean
        denominator += 4.0
        return numerator / denominator

    def update(self, budget: int, family: str, confidence: float, roi: float) -> None:
        bucket = self.bucket(confidence)
        self.global_stats[budget][bucket].update(roi)
        self.family_stats[budget][family][bucket].update(roi)
        self.budget_overall_stats[budget].update(roi)
        self.family_overall_stats[budget][family].update(roi)


@dataclass
class BudgetSummary:
    budget: int
    total_cost: int = 0
    total_return: int = 0
    total_profit: int = 0
    steps: int = 0
    prize_counts: Counter = field(default_factory=Counter)
    family_usage: Counter = field(default_factory=Counter)
    max_red_hits_seen: int = 0
    total_blue_hit_tickets: int = 0
    match_score_sum: float = 0.0
    winning_issue_count: int = 0
    value_hit_issue_count: int = 0
    strong_hit_issue_count: int = 0

    def add(self, family: str, scheme: Scheme, result: SchemeResult) -> None:
        self.total_cost += scheme.cost
        self.total_return += result.total_return
        self.total_profit += result.profit
        self.steps += 1
        self.prize_counts.update(result.prize_counts)
        self.family_usage[family] += 1
        self.max_red_hits_seen = max(self.max_red_hits_seen, result.max_red_hits)
        self.total_blue_hit_tickets += result.blue_hit_tickets
        self.match_score_sum += result.average_match_score
        self.winning_issue_count += int(result.winning_ticket_count > 0)
        self.value_hit_issue_count += int(result.value_hit_tickets > 0)
        self.strong_hit_issue_count += int(result.strong_hit_tickets > 0)

    def to_record(self) -> dict[str, object]:
        avg_match_score = self.match_score_sum / self.steps if self.steps else 0.0
        roi = self.total_profit / self.total_cost if self.total_cost else 0.0
        winning_issue_rate = self.winning_issue_count / self.steps if self.steps else 0.0
        value_hit_issue_rate = self.value_hit_issue_count / self.steps if self.steps else 0.0
        strong_hit_issue_rate = self.strong_hit_issue_count / self.steps if self.steps else 0.0
        return {
            "budget": self.budget,
            "steps": self.steps,
            "total_cost": self.total_cost,
            "total_return_all_prizes": self.total_return,
            "total_profit_all_prizes": self.total_profit,
            "roi_all_prizes": round(roi, 4),
            "avg_match_score": round(avg_match_score, 4),
            "max_red_hits_seen": self.max_red_hits_seen,
            "blue_hit_tickets": self.total_blue_hit_tickets,
            "winning_issue_count": self.winning_issue_count,
            "winning_issue_rate": round(winning_issue_rate, 4),
            "value_hit_issue_count": self.value_hit_issue_count,
            "value_hit_issue_rate": round(value_hit_issue_rate, 4),
            "strong_hit_issue_count": self.strong_hit_issue_count,
            "strong_hit_issue_rate": round(strong_hit_issue_rate, 4),
            "family_usage": dict(self.family_usage),
            "prize_counts": dict(self.prize_counts),
        }


def merge_draws_with_official(
    draws: Sequence[Draw],
    official_records: Sequence[dict[str, object]],
) -> tuple[list[Draw], dict[str, object]]:
    official_by_issue: dict[int, Draw] = {}
    for record in official_records:
        issue = parse_int(record.get("code"))
        blue = parse_int(record.get("blue"))
        reds = tuple(sorted(parse_int(number) for number in str(record.get("red") or "").split(",") if parse_int(number)))
        if issue is None or blue is None or len(reds) != 6:
            continue
        official_by_issue[issue] = Draw(
            issue=issue,
            draw_date=str(record.get("date") or ""),
            reds=reds,
            blue=blue,
            official=build_official_metrics(record),
        )

    merged: list[Draw] = []
    mismatches: list[int] = []
    official_covered = 0
    seen_issues: set[int] = set()
    for draw in draws:
        official_draw = official_by_issue.get(draw.issue)
        seen_issues.add(draw.issue)
        if official_draw is None:
            merged.append(draw)
            continue
        if official_draw.reds != draw.reds or official_draw.blue != draw.blue:
            mismatches.append(draw.issue)
            merged.append(draw)
            continue
        official_covered += 1
        merged.append(
            Draw(
                issue=draw.issue,
                draw_date=official_draw.draw_date or draw.draw_date,
                reds=draw.reds,
                blue=draw.blue,
                official=official_draw.official,
            )
        )

    appended = 0
    for issue in sorted(official_by_issue):
        if issue in seen_issues:
            continue
        merged.append(official_by_issue[issue])
        appended += 1

    merged.sort(key=lambda draw: draw.issue)
    return merged, {
        "official_record_count": len(official_by_issue),
        "official_coverage_on_excel": official_covered,
        "official_coverage_ratio_on_excel": round(official_covered / len(draws), 6) if draws else 0.0,
        "appended_latest_issue_count": appended,
        "latest_issue_after_merge": merged[-1].issue if merged else None,
        "mismatch_issues": mismatches,
    }


def scaled_log(value: int | None, ceiling: int) -> float:
    if value is None or value <= 0:
        return 0.0
    return min(math.log1p(value) / math.log1p(ceiling), 1.0)


def scaled_ratio(value: float | int | None, ceiling: float) -> float:
    if value is None or ceiling <= 0:
        return 0.0
    return min(max(float(value), 0.0) / ceiling, 1.0)


def average(values: Sequence[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def stddev(values: Sequence[float]) -> float:
    if len(values) <= 1:
        return 0.0
    mu = average(values)
    return math.sqrt(sum((value - mu) ** 2 for value in values) / len(values))


def load_draws(excel_path: Path) -> list[Draw]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    draws: list[Draw] = []
    for index in range(2, len(rows), 7):
        block = rows[index : index + 7]
        if len(block) < 7:
            continue
        issue = block[0][1]
        draw_date = block[0][2]
        if issue is None:
            continue
        numbers = [int(row[3]) for row in block]
        draws.append(
            Draw(
                issue=int(issue),
                draw_date=str(draw_date),
                reds=tuple(sorted(numbers[:6])),
                blue=int(numbers[6]),
            )
        )
    return sorted(draws, key=lambda draw: draw.issue)


def extract_rule_text(rule_docx: Path) -> str:
    with ZipFile(rule_docx) as archive:
        xml = archive.read("word/document.xml").decode("utf-8")
    return "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", xml))


def infer_next_issue(draws: Sequence[Draw]) -> int:
    last = draws[-1].issue
    year = last // 1000
    seq = last % 1000
    # 双色球每年约150期（周二/周四/周日），若当年期号达到上限则跨年
    if seq >= 153:
        return (year + 1) * 1000 + 1
    return last + 1


def infer_next_draw_date(last_draw_date: str) -> str:
    base_date = datetime.strptime(last_draw_date.split("(")[0], "%Y-%m-%d")
    for offset in range(1, 5):
        candidate = base_date + timedelta(days=offset)
        if candidate.weekday() in {1, 3, 6}:
            return candidate.strftime("%Y-%m-%d")
    return (base_date + timedelta(days=3)).strftime("%Y-%m-%d")


def normalize_scores(raw: dict[int, float]) -> dict[int, float]:
    if not raw:
        return {}
    low = min(raw.values())
    high = max(raw.values())
    if math.isclose(low, high):
        return {key: 0.5 for key in raw}
    return {key: (value - low) / (high - low) for key, value in raw.items()}


# ══════════════════════════════════════════════════════════════════════════════
# 负二项分布缺席建模（Negative Binomial Gap Modeling）
# 理论依据：每个号码两次出现之间的间隔期（gap）服从负二项分布 NegBin(r, p)
#           比简单的 log(gap) 更精确地估计"该号码当前是否该出现了"
# ══════════════════════════════════════════════════════════════════════════════

def _negbinom_hazard(k: int, r: float, p: float) -> float:
    """
    计算 NegBin(r,p) 在缺席期 k 处的危险率：h(k) = P(X=k) / P(X≥k)
    即：在已知号码已缺席 k 期的条件下，本期出现的概率。
    使用 log-space 递推防止数值溢出，最多计算到 k+1 项。
    参数化：P(X=k) = C(k+r-1,k) * p^r * (1-p)^k  (failures before r successes)
    """
    k = max(0, int(k))
    log_p   = math.log(max(p, 1e-10))
    log_1mp = math.log(max(1.0 - p, 1e-10))
    # 递推：log P(X=0) = r*log(p)；log P(X=j) = log P(X=j-1) + log(1-p) + log(j+r-1) - log(j)
    log_pmf = r * log_p
    pmf_at_k = 0.0
    cdf_before_k = 0.0
    for j in range(min(k + 1, 200)):
        pmf_j = math.exp(log_pmf)
        if j < k:
            cdf_before_k += pmf_j
        else:
            pmf_at_k = pmf_j
        if j < 199:
            log_pmf += log_1mp + math.log(j + r) - math.log(j + 1)
    survival = max(1.0 - cdf_before_k, 1e-10)
    return min(pmf_at_k / survival, 1.0)


def _negbinom_gap_score(gap_list: list[int], current_gap: int) -> float:
    """
    拟合负二项分布并计算当前缺席期的危险率，替代原来的 math.log1p(gap)。
    数据不足时（<4个间隔样本）回退到归一化对数缩放。
    """
    fallback = min(math.log1p(current_gap) / math.log1p(30), 1.0)
    if len(gap_list) < 4:
        return fallback
    mean_g = sum(gap_list) / len(gap_list)
    if mean_g <= 0:
        return fallback
    var_g = sum((g - mean_g) ** 2 for g in gap_list) / max(len(gap_list) - 1, 1)
    if var_g <= 0:
        return fallback
    if var_g <= mean_g:
        # 欠散布（或正好符合几何分布）：hazard 为常数（几何分布无记忆性）
        return min(max(1.0 / (1.0 + mean_g), 0.0), 1.0)
    # 过散布：矩法估计 NegBin(r, p)
    # mean = r(1-p)/p → p = mean/var；r = mean² / (var - mean)
    p_hat = min(max(mean_g / var_g, 0.01), 0.99)
    r_hat = max(mean_g * p_hat / (1.0 - p_hat), 0.1)
    return _negbinom_hazard(current_gap, r_hat, p_hat)


def compute_number_features(
    history: Sequence[Draw],
    max_number: int,
    hit_getter: Callable[[Draw, int], bool],
    decay_base: float,
    similarity_scores: dict[int, float] | None = None,
    markov_scores: dict[int, float] | None = None,
    entropy_scores: dict[int, float] | None = None,
    recurrence_scores: dict[int, float] | None = None,
) -> dict[str, dict[int, float]]:
    raw: dict[str, dict[int, float]] = {name: {} for name in FEATURE_NAMES}
    history_len = len(history)
    reversed_history = list(reversed(history))
    micro_window = min(5, max(history_len, 1))
    short_window = min(10, max(history_len, 1))
    mid_window = min(30, max(history_len, 1))
    long_window = min(60, max(history_len, 1))
    xl_window = min(100, max(history_len, 1))
    wide_window = min(180, max(history_len, 1))
    for number in range(1, max_number + 1):
        micro_hits = 0
        short_hits = 0
        mid_hits = 0
        long_hits = 0
        xl_hits = 0
        wide_hits = 0
        decay_score = 0.0
        gap = history_len
        seen = False
        gap_list: list[int] = []   # 负二项分布：收集历史缺席期序列
        _last_age: int | None = None  # 上次出现的 age（用于计算连续间隔）
        for age, draw in enumerate(reversed_history, start=1):
            hit = hit_getter(draw, number)
            if hit:
                if age <= 5:
                    micro_hits += 1
                if age <= 10:
                    short_hits += 1
                if age <= 30:
                    mid_hits += 1
                if age <= 60:
                    long_hits += 1
                if age <= 100:
                    xl_hits += 1
                if age <= 180:
                    wide_hits += 1
                decay_score += decay_base ** age
                # 负二项：记录相邻两次出现之间的缺席期数
                if _last_age is not None:
                    gap_list.append(age - _last_age - 1)
                _last_age = age
                if not seen:
                    gap = age - 1
                    seen = True
        micro_rate = micro_hits / micro_window
        short_rate = short_hits / short_window
        mid_rate = mid_hits / mid_window
        long_rate = long_hits / long_window
        xl_rate = xl_hits / xl_window
        wide_rate = wide_hits / wide_window
        trend = short_rate - mid_rate
        momentum = micro_rate - long_rate
        rebound = (math.log1p(gap) ** 1.25) * max(0.0, long_rate - short_rate + 0.08)
        consistency = 1.0 - abs(short_rate - long_rate)
        raw["micro"][number] = micro_rate
        raw["short"][number] = short_rate
        raw["mid"][number] = mid_rate
        raw["long"][number] = long_rate
        raw["xl"][number] = xl_rate
        raw["wide"][number] = wide_rate
        raw["decay"][number] = decay_score
        # 负二项分布危险率：h(k) = P(今期出现 | 已缺席k期)，替代原 log1p(gap)
        raw["gap"][number] = _negbinom_gap_score(gap_list, gap)
        raw["trend"][number] = trend
        raw["momentum"][number] = momentum
        raw["rebound"][number] = rebound
        raw["consistency"][number] = consistency
        raw["similarity"][number] = 0.0 if similarity_scores is None else similarity_scores.get(number, 0.0)
        # ── 混沌 & 概率论新增特征 ────────────────────────────────────────────
        raw["entropy"][number]    = 0.0 if entropy_scores    is None else entropy_scores.get(number, 0.0)
        raw["markov"][number]     = 0.0 if markov_scores     is None else markov_scores.get(number, 0.0)
        raw["recurrence"][number] = 0.0 if recurrence_scores is None else recurrence_scores.get(number, 0.0)
    return {name: normalize_scores(values) for name, values in raw.items()}


def score_arm_numbers(
    feature_maps: dict[str, dict[int, float]],
    arm_weights: dict[str, float],
) -> dict[int, float]:
    raw_scores: dict[int, float] = {}
    numbers = next(iter(feature_maps.values())).keys()
    for number in numbers:
        raw_scores[number] = sum(
            arm_weights.get(feature_name, 0.0) * feature_maps[feature_name][number]
            for feature_name in feature_maps
        )
    return normalize_scores(raw_scores)


def rank_map(scores: dict[int, float]) -> dict[int, int]:
    ranked = sorted(scores, key=lambda number: (-scores[number], number))
    return {number: index + 1 for index, number in enumerate(ranked)}


def red_score_quality(scores: dict[int, float], actual_reds: Sequence[int]) -> float:
    ranks = rank_map(scores)
    rank_quality = sum((RED_MAX + 1 - ranks[number]) / RED_MAX for number in actual_reds) / len(actual_reds)
    actual_mask = numbers_to_mask(actual_reds)
    top_8_cover = (numbers_to_mask(sorted(scores, key=scores.get, reverse=True)[:8]) & actual_mask).bit_count() / len(actual_reds)
    top_13_cover = (numbers_to_mask(sorted(scores, key=scores.get, reverse=True)[:13]) & actual_mask).bit_count() / len(actual_reds)
    actual_score = sum(scores[number] for number in actual_reds) / len(actual_reds)
    return 0.35 * rank_quality + 0.25 * actual_score + 0.20 * top_8_cover + 0.20 * top_13_cover


def blue_score_quality(scores: dict[int, float], actual_blue: int) -> float:
    ranks = rank_map(scores)
    rank_quality = (BLUE_MAX + 1 - ranks[actual_blue]) / BLUE_MAX
    actual_score = scores[actual_blue]
    sorted_blues = sorted(scores, key=scores.get, reverse=True)
    top_3_hit = 1.0 if actual_blue in set(sorted_blues[:3]) else 0.0
    top_6_hit = 0.5 if actual_blue in set(sorted_blues[:6]) else 0.0
    return 0.40 * rank_quality + 0.32 * actual_score + 0.18 * top_3_hit + 0.10 * top_6_hit


def update_weights(weights: dict[str, float], qualities: dict[str, float], eta: float = 0.65) -> dict[str, float]:
    average_quality = sum(qualities.values()) / len(qualities)
    n = len(weights)
    # 最小权重下限提高：0.06（旧:0.04），防止cold_rebound等策略被压制到不可用
    # 2026032教训：cold_rebound仅13.8%导致冷号11/31完全漏判
    min_weight = max(0.06, 0.5 / n)
    updated: dict[str, float] = {}
    for arm, value in weights.items():
        updated[arm] = max(min_weight, value * math.exp(eta * (qualities[arm] - average_quality)))
    total = sum(updated.values())
    return {arm: value / total for arm, value in updated.items()}


def blend_score_maps(source_weights: dict[str, float], source_maps: dict[str, dict[int, float]]) -> dict[int, float]:
    numbers = next(iter(source_maps.values())).keys()
    raw_scores = {
        number: sum(source_weights[source] * source_maps[source][number] for source in source_maps)
        for number in numbers
    }
    return normalize_scores(raw_scores)


def build_history_context_vector(history: Sequence[Draw]) -> np.ndarray:
    recent10 = history[-10:] if len(history) >= 10 else history
    recent30 = history[-30:] if len(history) >= 30 else history
    last_draw = history[-1]
    previous_draw = history[-2] if len(history) >= 2 else history[-1]
    sum_scale = 33 * 6
    blue_values = [draw.blue for draw in recent10]
    repeat_values = []
    for left, right in zip(recent30, recent30[1:]):
        repeat_values.append((left.red_mask & right.red_mask).bit_count() / 6.0)
    official_recent10 = [draw.official for draw in recent10 if draw.official is not None]
    official_recent30 = [draw.official for draw in recent30 if draw.official is not None]
    sales10 = [scaled_log(metrics.sales, 600_000_000) for metrics in official_recent10]
    sales30 = [scaled_log(metrics.sales, 600_000_000) for metrics in official_recent30]
    pool10 = [scaled_log(metrics.poolmoney, 3_200_000_000) for metrics in official_recent10]
    pool30 = [scaled_log(metrics.poolmoney, 3_200_000_000) for metrics in official_recent30]
    first_money30 = [scaled_log(metrics.prize_amount(1), 20_000_000) for metrics in official_recent30]
    first_count30 = [scaled_ratio(metrics.prize_count(1), 20.0) for metrics in official_recent30]
    second_count30 = [scaled_ratio(metrics.prize_count(2), 300.0) for metrics in official_recent30]
    region_count30 = [scaled_ratio(metrics.first_region_count, 12.0) for metrics in official_recent30]
    region_entropy30 = [metrics.first_region_entropy for metrics in official_recent30]
    fixed_burden30 = [
        min(metrics.fixed_prize_total / metrics.sales, 0.25)
        for metrics in official_recent30
        if metrics.sales
    ]
    last_sales = scaled_log(last_draw.official.sales, 600_000_000) if last_draw.official else 0.0
    last_pool = scaled_log(last_draw.official.poolmoney, 3_200_000_000) if last_draw.official else 0.0
    sales_trend = (sales10[-1] - average(sales30)) if sales10 and sales30 else 0.0
    pool_trend = (pool10[-1] - average(pool30)) if pool10 and pool30 else 0.0
    vector = np.array(
        [
            sum(last_draw.reds) / sum_scale,
            (last_draw.reds[-1] - last_draw.reds[0]) / 32.0,
            sum(number % 2 for number in last_draw.reds) / 6.0,
            zone_split(last_draw.reds)[0] / 6.0,
            zone_split(last_draw.reds)[1] / 6.0,
            zone_split(last_draw.reds)[2] / 6.0,
            (last_draw.red_mask & previous_draw.red_mask).bit_count() / 6.0,
            sum(sum(draw.reds) for draw in recent10) / (len(recent10) * sum_scale),
            sum(count_consecutive_pairs(draw.reds) for draw in recent10) / max(len(recent10), 1),
            sum(sum(number % 2 for number in draw.reds) for draw in recent10) / (len(recent10) * 6.0),
            sum(draw.blue for draw in recent10) / (len(recent10) * BLUE_MAX),
            (float(np.std(blue_values)) / BLUE_MAX) if blue_values else 0.0,
            sum(sum(draw.reds) for draw in recent30) / (len(recent30) * sum_scale),
            (sum(repeat_values) / len(repeat_values)) if repeat_values else 0.0,
            last_sales,
            average(sales10),
            sales_trend,
            min(stddev(sales30) / 0.08, 1.0) if sales30 else 0.0,
            last_pool,
            average(pool10),
            pool_trend,
            min(stddev(pool30) / 0.10, 1.0) if pool30 else 0.0,
            average(first_money30),
            average(first_count30),
            average(second_count30),
            average(region_count30),
            average(region_entropy30),
            min(average(fixed_burden30) / 0.12, 1.0) if fixed_burden30 else 0.0,
        ],
        dtype=float,
    )
    return vector


def cosine_similarity(a: np.ndarray, b: np.ndarray) -> float:
    norm_a = float(np.linalg.norm(a))
    norm_b = float(np.linalg.norm(b))
    if norm_a < 1e-9 or norm_b < 1e-9:
        return 0.0
    return float(np.dot(a, b)) / (norm_a * norm_b)


def build_similarity_feature_map(
    current_context: np.ndarray,
    snapshots: Sequence[HistorySnapshot],
    max_number: int,
    number_getter: Callable[[HistorySnapshot], Sequence[int]],
    top_k: int,
) -> dict[int, float]:
    if not snapshots:
        return {number: 0.0 for number in range(1, max_number + 1)}
    similarities: list[tuple[float, HistorySnapshot]] = []
    total_snapshots = len(snapshots)
    for index, snapshot in enumerate(snapshots):
        # 结合欧氏距离和余弦相似度：欧氏距离关注绝对差异，余弦关注方向相似性
        euclidean_dist = float(np.linalg.norm(current_context - snapshot.context_vector))
        cosine_sim = cosine_similarity(current_context, snapshot.context_vector)
        recency_weight = 0.985 ** (total_snapshots - index - 1)
        # 融合两种相似度：欧氏 60% + 余弦 40%
        euclidean_sim = math.exp(-3.2 * euclidean_dist)
        cosine_component = (cosine_sim + 1.0) / 2.0  # 归一化到 [0, 1]
        similarity = (0.60 * euclidean_sim + 0.40 * cosine_component) * recency_weight
        similarities.append((similarity, snapshot))
    top_matches = sorted(similarities, key=lambda item: item[0], reverse=True)[:top_k]
    raw_scores = {number: 0.0 for number in range(1, max_number + 1)}
    for similarity, snapshot in top_matches:
        if similarity <= 0.0:
            continue
        for number in number_getter(snapshot):
            raw_scores[number] += similarity
    return normalize_scores(raw_scores)


def zone_index(number: int) -> int:
    if number <= 11:
        return 0
    if number <= 22:
        return 1
    return 2


def zone_split(numbers: Sequence[int]) -> tuple[int, int, int]:
    counts = [0, 0, 0]
    for number in numbers:
        counts[zone_index(number)] += 1
    return tuple(counts)


def count_consecutive_pairs(numbers: Sequence[int]) -> int:
    return sum(1 for left, right in zip(numbers, numbers[1:]) if right - left == 1)


def smooth_counter_probability(counter: Counter, key: object) -> float:
    return (counter.get(key, 0) + 1) / (sum(counter.values()) + len(counter) + 1)


def build_shape_context(history: Sequence[Draw]) -> dict[str, object]:
    sample = history[-200:] if len(history) > 200 else history
    odd_counts: Counter = Counter()
    zone_patterns: Counter = Counter()
    repeat_counts: Counter = Counter()
    consecutive_counts: Counter = Counter()
    sums: list[int] = []
    spans: list[int] = []
    previous_draw: Draw | None = None
    for draw in sample:
        odd_counts[sum(number % 2 for number in draw.reds)] += 1
        zone_patterns[zone_split(draw.reds)] += 1
        consecutive_counts[count_consecutive_pairs(draw.reds)] += 1
        sums.append(sum(draw.reds))
        spans.append(draw.reds[-1] - draw.reds[0])
        if previous_draw is not None:
            repeat_counts[(draw.red_mask & previous_draw.red_mask).bit_count()] += 1
        previous_draw = draw
    # ── 持续性掩码：近2期出现过的号码集合（用于 persistence_score）
    last_2_mask = history[-1].red_mask
    if len(history) >= 2:
        last_2_mask |= history[-2].red_mask
    return {
        "odd_counts": odd_counts,
        "zone_patterns": zone_patterns,
        "repeat_counts": repeat_counts,
        "consecutive_counts": consecutive_counts,
        "sum_mean": sum(sums) / len(sums),
        "sum_std": max(6.0, math.sqrt(sum((value - (sum(sums) / len(sums))) ** 2 for value in sums) / len(sums))),
        "span_mean": sum(spans) / len(spans),
        "last_reds": tuple(history[-1].reds),
        "last_red_mask": history[-1].red_mask,
        "last_2_red_mask": last_2_mask,  # 近2期红球并集掩码
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 混沌原理 & 概率论模块
# 理论基础：
#   1. Shannon熵  H = -Σ p_i·log₂(p_i)  — 测量分布的不确定性
#      当某号码近期出现率偏离均匀分布（偏离最大熵 log₂N），说明进入"可预测态"
#      即混沌系统中的"吸引子盆地"，短期内更容易预测
#   2. 一阶Markov链  P(X_t=i | X_{t-1}=j)  — 条件转移概率
#      利用历史相邻两期共现频率估计，Laplace平滑处理小样本问题
#   3. 递归定量分析 (RQA)  — 统计号码在1~12期滞后中的周期性递归
#      识别双色球开奖序列中的节律性和准周期结构
#   4. 相位空间重构 (Takens定理) — 注入red_combo_score
#      将历史开奖和数序列嵌入高维相位空间，找"吸引子"密集区域
# ═══════════════════════════════════════════════════════════════════════════════

def build_markov_transition(
    history: Sequence[Draw],
    max_number: int,
    window: int = 120,
    decay: float = 0.980,
) -> dict[int, dict[int, float]]:
    """
    构建一阶Markov转移矩阵（带指数衰减）。
    trans[i][j] = P(号码j在t期出现 | 号码i在t-1期出现)
    使用Laplace平滑：(count + 1) / (total + max_number)
    """
    # count[i][j] = i出现后j出现的加权频次
    count: dict[int, dict[int, float]] = {
        i: defaultdict(float) for i in range(1, max_number + 1)
    }
    sample = history[-window:] if len(history) > window else history
    total = len(sample)
    for age, (prev_draw, curr_draw) in enumerate(
        zip(reversed(sample[:-1]), reversed(sample[1:])), start=1
    ):
        w = decay ** (total - age - 1)
        for prev_num in prev_draw.reds:
            if prev_num > max_number:
                continue
            for curr_num in curr_draw.reds:
                if curr_num > max_number:
                    continue
                count[prev_num][curr_num] += w
    # Laplace平滑归一化
    trans: dict[int, dict[int, float]] = {}
    for i in range(1, max_number + 1):
        row_total = sum(count[i].values()) + max_number  # +N for Laplace
        trans[i] = {
            j: (count[i].get(j, 0.0) + 1.0) / row_total
            for j in range(1, max_number + 1)
        }
    return trans


def build_markov_scores(
    history: Sequence[Draw],
    max_number: int,
    markov_trans: dict[int, dict[int, float]],
) -> dict[int, float]:
    """
    根据Markov矩阵计算每个号码的"转入概率"：
    score(j) = Σ_{i in last_draw} P(j | i)
    即上期所有出现号码对本期各号码的条件概率之和。
    """
    if not history:
        return {n: 0.0 for n in range(1, max_number + 1)}
    last_draw = history[-1]
    scores: dict[int, float] = {}
    for j in range(1, max_number + 1):
        scores[j] = sum(
            markov_trans[i].get(j, 0.0)
            for i in last_draw.reds
            if i <= max_number
        )
    return scores


def build_shannon_entropy_scores(
    history: Sequence[Draw],
    max_number: int,
    hit_getter: Callable[[Draw, int], bool],
    windows: tuple[int, ...] = (10, 30, 60),
) -> dict[int, float]:
    """
    Shannon熵驱动的可预测性评分。
    对每个号码，在多窗口内计算出现率分布的熵偏差：
      entropy_deviation = log₂(window) - H(p_hit, p_miss)
    偏差越大 → 分布越偏斜 → 越偏离随机态 → 越接近"混沌吸引子"可预测区
    取多窗口加权平均（短期权重更高）。
    """
    log2 = math.log2
    raw: dict[int, float] = {}
    for number in range(1, max_number + 1):
        total_dev = 0.0
        total_w = 0.0
        for win in windows:
            w = 1.0 / win  # 短窗口权重更高
            sample = history[-win:] if len(history) >= win else history
            n = len(sample)
            if n == 0:
                continue
            hits = sum(1 for draw in sample if hit_getter(draw, number))
            p = hits / n
            q = 1.0 - p
            # 二元Shannon熵
            if p <= 0.0 or p >= 1.0:
                h = 0.0  # 完全确定，最大偏差
            else:
                h = -(p * log2(p) + q * log2(q))
            max_h = 1.0  # log₂(2) = 1
            dev = max_h - h  # 0=最大熵(均匀), 1=完全确定
            total_dev += w * dev
            total_w += w
        raw[number] = total_dev / total_w if total_w > 0 else 0.0
    return raw


def build_recurrence_scores(
    history: Sequence[Draw],
    max_number: int,
    hit_getter: Callable[[Draw, int], bool],
    max_lag: int = 12,
    decay: float = 0.90,
) -> dict[int, float]:
    """
    递归定量分析（RQA简化版）。
    对每个号码，统计它在当前期的1~max_lag期之前是否也出现过，
    并用指数衰减加权求和，捕捉周期性节律（如每3期/每7期出现的模式）。
    recurrence(n) = Σ_{lag=1}^{max_lag} decay^(lag-1) · hit(n, t-lag)
    """
    raw: dict[int, float] = {}
    if not history:
        return {n: 0.0 for n in range(1, max_number + 1)}
    for number in range(1, max_number + 1):
        score = 0.0
        for lag in range(1, min(max_lag + 1, len(history) + 1)):
            draw = history[-lag]
            if hit_getter(draw, number):
                score += decay ** (lag - 1)
        raw[number] = score
    return raw


def build_phase_attractor_score(
    history: Sequence[Draw],
    embed_dim: int = 3,
    lag: int = 1,
    top_k: int = 20,
) -> dict[str, float]:
    """
    Takens嵌入定理：相位空间重构。
    将历史开奖红球和数序列嵌入 embed_dim 维相位空间，
    找到与当前状态最近邻的 top_k 个历史点，
    返回这些历史点的"后继状态"统计，用于预测下期号码分布。
    返回 {'sum_mean': float, 'sum_std': float, 'attractor_density': float}
    供 red_combo_score 中的 sum_score 校正使用。
    """
    if len(history) < embed_dim * lag + 1:
        return {"attractor_correction": 0.0, "attractor_density": 0.0}

    # 构建相位空间向量：每个点 = (sum_t, sum_{t-lag}, sum_{t-2*lag})
    def make_vector(draws: Sequence[Draw], idx: int) -> np.ndarray:
        return np.array([
            sum(draws[idx - k * lag].reds) / (33 * 6)
            for k in range(embed_dim)
        ])

    # 当前状态向量
    current_vec = make_vector(history, len(history) - 1)

    # 所有历史状态（可以有后继的点）
    candidates = []
    for i in range(embed_dim * lag, len(history) - 1):
        vec = make_vector(history, i)
        dist = float(np.linalg.norm(current_vec - vec))
        next_sum = sum(history[i + 1].reds)
        candidates.append((dist, next_sum))

    if not candidates:
        return {"attractor_correction": 0.0, "attractor_density": 0.0}

    candidates.sort(key=lambda x: x[0])
    top = candidates[:top_k]
    nearest_sums = [s for _, s in top]
    mean_sum = sum(nearest_sums) / len(nearest_sums)
    # 当前历史均值
    hist_mean = sum(sum(d.reds) for d in history[-60:]) / min(60, len(history)) if history else mean_sum
    # 吸引子修正：相位空间近邻预测的和数与历史均值的偏差
    correction = (mean_sum - hist_mean) / (33 * 6)
    # 密度：最近邻中距离很近的比例（衡量当前状态是否在吸引子稠密区）
    threshold = 0.08
    density = sum(1 for d, _ in top if d < threshold) / top_k
    return {"attractor_correction": correction, "attractor_density": density}


def build_pair_scores(history: Sequence[Draw], window: int = 100, decay: float = 0.982) -> tuple[dict[tuple[int, int], float], float]:
    """二元组共现分数：扩大窗口到100期(旧:80)，衰减略慢(旧:0.985)。"""
    pair_scores: dict[tuple[int, int], float] = defaultdict(float)
    for age, draw in enumerate(reversed(history[-window:]), start=1):
        weight = decay ** (age - 1)
        for pair in combinations(draw.reds, 2):
            pair_scores[pair] += weight
    max_pair_score = max(pair_scores.values(), default=1.0)
    return pair_scores, max_pair_score


def build_triplet_scores(history: Sequence[Draw], window: int = 150, decay: float = 0.978) -> tuple[dict[tuple[int, int, int], float], float]:
    """三元组共现分数：捕捉三个号码经常一起出现的模式。"""
    triplet_scores: dict[tuple[int, int, int], float] = defaultdict(float)
    for age, draw in enumerate(reversed(history[-window:]), start=1):
        weight = decay ** (age - 1)
        for triplet in combinations(draw.reds, 3):
            triplet_scores[triplet] += weight
    max_triplet_score = max(triplet_scores.values(), default=1.0)
    return triplet_scores, max_triplet_score


# ══════════════════════════════════════════════════════════════════════════════
# PageRank 共现网络中心性（Co-occurrence Network PageRank）
# 理论依据：把红球1~33视作图节点，两球在同一期出现则连边（权重=衰减频次）
#           PageRank 排名高 = 经常与其他"重要"球共现 = 下期更可能出现
# ══════════════════════════════════════════════════════════════════════════════

def build_pagerank_scores(
    history: Sequence[Draw],
    window: int = 150,
    decay: float = 0.982,
    alpha: float = 0.85,
    max_iter: int = 80,
    tol: float = 1e-6,
) -> dict[int, float]:
    """
    构建红球共现无向加权图，并用 PageRank 算法计算各号码的中心性得分。
    - window: 使用最近 window 期的历史
    - decay:  越近的期权重越高（与 build_pair_scores 保持一致）
    - alpha:  阻尼系数（标准值 0.85）
    返回：{ball_number: pagerank_score}（未归一化，供 red_combo_score 使用）
    """
    n = RED_MAX  # 33
    # 构建衰减加权共现矩阵 W (1-indexed, shape (n+1)×(n+1))
    W = np.zeros((n + 1, n + 1), dtype=float)
    for age, draw in enumerate(reversed(history[-window:]), start=1):
        w = decay ** (age - 1)
        for r1, r2 in combinations(draw.reds, 2):
            W[r1, r2] += w
            W[r2, r1] += w
    # 取 1~n 子矩阵，行归一化 → 随机转移矩阵 P
    M = W[1:, 1:]                                    # shape (33, 33)
    row_sums = M.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1.0                    # 避免除零（孤立节点均匀跳转）
    P = M / row_sums                                  # 行随机矩阵
    # Power iteration: r = α·Pᵀ·r + (1-α)/n
    r = np.full(n, 1.0 / n, dtype=float)
    for _ in range(max_iter):
        r_new = alpha * (P.T @ r) + (1.0 - alpha) / n
        if float(np.abs(r_new - r).sum()) < tol:
            r = r_new
            break
        r = r_new
    return {i + 1: float(r[i]) for i in range(n)}


def red_combo_score(
    reds: tuple[int, ...],
    red_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
) -> float:
    # 三元组数据通过 shape_context 注入（若调用者已填充 "triplet_scores" 键）
    triplet_scores = shape_context.get("triplet_scores")
    max_triplet_score = float(shape_context.get("max_triplet_score", 1.0))
    base_score = sum(red_scores[number] for number in reds) / len(reds)
    pair_score = (
        sum(pair_scores.get(tuple(sorted(pair)), 0.0) for pair in combinations(reds, 2))
        / (15 * max_pair_score)
        if max_pair_score
        else 0.0
    )
    # 三元组共现评分：捕捉三个号码历史上经常一起出现的模式
    triplet_score = 0.0
    if triplet_scores is not None and max_triplet_score > 0:
        triplet_score = (
            sum(triplet_scores.get(tuple(sorted(t)), 0.0) for t in combinations(reds, 3))
            / (20 * max_triplet_score)
        )
    odd_prob = smooth_counter_probability(shape_context["odd_counts"], sum(number % 2 for number in reds))
    zone_prob = smooth_counter_probability(shape_context["zone_patterns"], zone_split(reds))
    repeat_prob = smooth_counter_probability(shape_context["repeat_counts"], (numbers_to_mask(reds) & shape_context["last_red_mask"]).bit_count())
    consecutive_prob = smooth_counter_probability(shape_context["consecutive_counts"], count_consecutive_pairs(reds))
    # ── PageRank 共现网络中心性（组合平均PageRank得分）─────────────────────
    # 当该组合中每个球的共现中心性都高时，整体得分高
    _pr_scores = shape_context.get("pagerank_scores")
    _pr_max    = float(shape_context.get("pagerank_max", 1.0))
    pagerank_combo = (
        sum(_pr_scores.get(r, 0.0) for r in reds) / (len(reds) * _pr_max)
        if _pr_scores and _pr_max > 0 else 0.0
    )
    # ── 持续性评分：候选红球中有多少个出现在近2期开奖中 ──────────────────
    # 2026032实测：33连续出现在031/032，但旧权重(repeat_prob=0.06)严重低估此信号
    last_2_mask = shape_context.get("last_2_red_mask", shape_context["last_red_mask"])
    persistence_hits = (numbers_to_mask(reds) & last_2_mask).bit_count()
    persistence_score = min(persistence_hits / 3.0, 1.0)  # 0~2个归一化
    # ── 相位空间吸引子修正（Takens嵌入） ────────────────────────────────────
    attractor_correction = float(shape_context.get("attractor_correction", 0.0))
    attractor_density    = float(shape_context.get("attractor_density", 0.0))
    adjusted_sum_mean = shape_context["sum_mean"] + attractor_correction * shape_context["sum_std"]
    sum_score = math.exp(-((sum(reds) - adjusted_sum_mean) / shape_context["sum_std"]) ** 2 / 2)
    attractor_bonus = attractor_density * sum_score
    span_score = 1.0 - min(abs((reds[-1] - reds[0]) - shape_context["span_mean"]) / 25.0, 1.0)
    return (
        0.27 * base_score          # 单号评分（旧:0.33 → 让出权重给PageRank）
        + 0.06 * pagerank_combo    # 【新】PageRank共现网络中心性（从base_score分出）
        + 0.11 * pair_score        # 二元组共现
        + 0.08 * triplet_score     # 三元组共现
        + 0.07 * odd_prob          # 奇偶分布
        + 0.07 * zone_prob         # 区间分布
        + 0.09 * repeat_prob       # 上期重复数分布匹配
        + 0.04 * consecutive_prob  # 连号
        + 0.05 * persistence_score # 近2期持续性奖励
        + 0.05 * sum_score         # 和数高斯核
        + 0.04 * attractor_bonus   # 相位空间吸引子奖励
        + 0.04 * span_score        # 跨度匹配
        + 0.01 * attractor_density # 稠密区额外加分
        # weights sum = 0.27+0.06+0.11+0.08+0.07+0.07+0.09+0.04+0.05+0.05+0.04+0.04+0.01 = 0.98
        # remaining 0.02 acts as implicit regularization
    )


def ticket_score(
    reds: tuple[int, ...],
    blue: int,
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    red_cache: dict[tuple[int, ...], float],
    triplet_scores: dict[tuple[int, int, int], float] | None = None,
    max_triplet_score: float = 1.0,
) -> float:
    if reds not in red_cache:
        red_cache[reds] = red_combo_score(
            reds, red_scores, shape_context, pair_scores, max_pair_score,
        )
    return 0.86 * red_cache[reds] + 0.14 * blue_scores[blue]


def normalized_value_map(values: dict[str, float]) -> dict[str, float]:
    return normalize_scores(values) if values else {}


def family_selection_score(stat_score: float, proxy_score: float, count: int) -> float:
    exploration = 0.10 / math.sqrt(count + 1)
    return 0.48 * stat_score + 0.42 * proxy_score + exploration


def budget_utilization_floor(budget: int) -> float:
    # 降低利用率下限：进场时不必用完预算，保留资金等待更强信号
    if budget <= 10:
        return 0.90   # 旧:0.95
    if budget <= 50:
        return 0.30   # 旧:0.45
    if budget <= 100:
        return 0.35   # 旧:0.50
    if budget <= 200:
        return 0.40   # 旧:0.55
    if budget <= 500:
        return 0.45   # 旧:0.60
    return 0.50       # 旧:0.65


def gate_hold_threshold(budget: int) -> float:
    # 极端严格：只有当校准器预测ROI明显好于先验(-0.45)时才入场
    # 目标：将无效入场率降低到<10%，接近理论最优ROI上限
    if budget <= 10:
        return 0.10   # 预测ROI需>10%才入场(旧:-0.20)
    if budget <= 50:
        return 0.15   # 旧:-0.12
    if budget <= 100:
        return 0.18   # 旧:-0.06
    if budget <= 200:
        return 0.20   # 旧:-0.02
    if budget <= 500:
        return 0.22   # 旧:0.00
    return 0.25       # 旧:0.04


def scheme_gate_confidence(
    family_stat_score: float,
    family_proxy_score: float,
    scheme: Scheme,
    red_similarity_peak: float,
    blue_similarity_peak: float,
) -> float:
    budget = int(scheme.metadata.get("budget", scheme.cost)) if scheme.metadata else scheme.cost
    utilization = scheme.cost / budget if budget else 0.0
    # 增大相似期信号权重：当前局面与历史高回报局面越相似，信心越高
    confidence = (
        0.42 * family_stat_score
        + 0.10 * family_proxy_score
        + 0.06 * utilization
        + 0.28 * red_similarity_peak   # 红球相似度权重大幅提升(旧:0.10)
        + 0.14 * blue_similarity_peak  # 蓝球相似度权重大幅提升(旧:0.06)
    )
    return min(max(confidence, 0.0), 1.0)


def make_hold_scheme(budget: int, predicted_gate_roi: float) -> Scheme:
    return Scheme(
        family="hold",
        description="空仓观望",
        cost=0,
        tickets=[],
        proxy_score=predicted_gate_roi,
        metadata={
            "budget": budget,
            "predicted_gate_roi": round(predicted_gate_roi, 6),
            "decision": "hold",
        },
    )


NESTED_VALIDATION_WINDOWS = (12, 24, 48)
NESTED_VALIDATION_MIN_OBSERVATIONS = 32   # 需要更多历史观测(旧:24)
NESTED_VALIDATION_MIN_AVAILABLE_WINDOWS = 2
NESTED_VALIDATION_MIN_PASS_RATIO = 3 / 4  # 需要75%窗口正ROI(旧:2/3)


@dataclass(frozen=True)
class NestedValidationSummary:
    passed: bool
    score: float
    observations: int
    overall_roi: float
    recent_roi: float
    pass_ratio: float
    overall_profitable_hits: int
    recent_profitable_hits: int
    positive_return_concentration: float
    window_rois: tuple[tuple[int, float], ...]
    reason: str


@dataclass(frozen=True)
class HitValidationSummary:
    passed: bool
    score: float
    observations: int
    overall_hit_score: float
    recent_hit_score: float
    overall_any_hit_rate: float
    recent_any_hit_rate: float
    overall_value_hit_rate: float
    recent_value_hit_rate: float
    pass_ratio: float
    window_hit_scores: tuple[tuple[int, float], ...]
    reason: str
    overall_strong_hit_rate: float = 0.0  # 四等奖及以上命中率


@dataclass(frozen=True)
class ProfileValidationSummary:
    passed: bool
    score: float
    mode: str
    reason: str
    roi: NestedValidationSummary
    hit: HitValidationSummary


def validation_roi_score(roi: float) -> float:
    # 提高标准：roi=-0.15时得0分(旧:-0.30)，roi=1.35时得满分(旧:0.60)
    # 使得ROI=-20%以上才能得到正向评分
    return min(max((roi + 0.15) / 1.50, 0.0), 1.0)


def compute_window_roi(records: Sequence[ValidationRecord], window: int | None = None) -> float:
    subset = records[-window:] if window is not None else records
    total_cost = sum(record.cost for record in subset)
    total_return = sum(record.total_return for record in subset)
    if total_cost <= 0:
        return 0.0
    return (total_return / total_cost) - 1.0


def count_profitable_hits(records: Sequence[ValidationRecord], window: int | None = None) -> int:
    subset = records[-window:] if window is not None else records
    return sum(1 for record in subset if record.total_return > record.cost)


def positive_return_concentration(records: Sequence[ValidationRecord]) -> float:
    positive_returns = [record.total_return for record in records if record.total_return > 0]
    if not positive_returns:
        return 1.0
    total_positive_return = sum(positive_returns)
    if total_positive_return <= 0:
        return 1.0
    return max(positive_returns) / total_positive_return


def compute_window_hit_score(records: Sequence[ValidationRecord], window: int | None = None) -> float:
    subset = records[-window:] if window is not None else records
    if not subset:
        return 0.0
    return sum(record.hit_score for record in subset) / len(subset)


def compute_window_any_hit_rate(records: Sequence[ValidationRecord], window: int | None = None) -> float:
    subset = records[-window:] if window is not None else records
    if not subset:
        return 0.0
    return sum(1.0 for record in subset if record.any_prize_hit) / len(subset)


def compute_window_value_hit_rate(records: Sequence[ValidationRecord], window: int | None = None) -> float:
    subset = records[-window:] if window is not None else records
    if not subset:
        return 0.0
    return sum(1.0 for record in subset if record.value_hit) / len(subset)


def compute_window_strong_hit_rate(records: Sequence[ValidationRecord], window: int | None = None) -> float:
    """四等奖及以上（200元+）命中率，防止方案永远只命中六等奖(5元)。"""
    subset = records[-window:] if window is not None else records
    if not subset:
        return 0.0
    return sum(1.0 for record in subset if record.strong_hit) / len(subset)


def hit_validation_thresholds(budget: int) -> tuple[float, float, float, float]:
    # 提高命中质量要求：只有显著强于随机的profile才能通过
    # (overall_score, recent_score, any_hit_rate, recent_any_hit_rate)
    if budget <= 10:
        return 0.24, 0.22, 0.22, 0.20   # 旧:(0.18,0.16,0.18,0.16)
    if budget <= 50:
        return 0.30, 0.28, 0.38, 0.35   # 旧:(0.22,0.20,0.30,0.28)
    if budget <= 100:
        return 0.32, 0.30, 0.42, 0.38   # 旧:(0.24,0.22,0.34,0.30)
    if budget <= 200:
        return 0.34, 0.32, 0.46, 0.42   # 旧:(0.25,0.23,0.36,0.32)
    if budget <= 500:
        return 0.36, 0.33, 0.52, 0.46   # 旧:(0.27,0.24,0.42,0.36)
    return 0.38, 0.35, 0.58, 0.52       # 旧:(0.28,0.25,0.48,0.40)


def summarize_nested_validation(records: Sequence[ValidationRecord]) -> NestedValidationSummary:
    observations = len(records)
    if observations < NESTED_VALIDATION_MIN_OBSERVATIONS:
        return NestedValidationSummary(
            passed=False,
            score=0.0,
            observations=observations,
            overall_roi=0.0,
            recent_roi=0.0,
            pass_ratio=0.0,
            overall_profitable_hits=0,
            recent_profitable_hits=0,
            positive_return_concentration=1.0,
            window_rois=(),
            reason="insufficient_history",
        )
    available_window_rois: list[tuple[int, float]] = []
    for window in NESTED_VALIDATION_WINDOWS:
        if observations < max(window, NESTED_VALIDATION_MIN_OBSERVATIONS):
            continue
        available_window_rois.append((window, compute_window_roi(records, window)))
    if len(available_window_rois) < NESTED_VALIDATION_MIN_AVAILABLE_WINDOWS:
        return NestedValidationSummary(
            passed=False,
            score=0.0,
            observations=observations,
            overall_roi=compute_window_roi(records),
            recent_roi=0.0,
            pass_ratio=0.0,
            overall_profitable_hits=count_profitable_hits(records),
            recent_profitable_hits=0,
            positive_return_concentration=positive_return_concentration(records),
            window_rois=tuple(available_window_rois),
            reason="insufficient_windows",
        )
    overall_roi = compute_window_roi(records)
    recent_roi = available_window_rois[0][1]
    positive_windows = sum(window_roi > 0.0 for _, window_roi in available_window_rois)
    pass_ratio = positive_windows / len(available_window_rois)
    sorted_rois = sorted(window_roi for _, window_roi in available_window_rois)
    median_roi = sorted_rois[len(sorted_rois) // 2]
    recent_window = available_window_rois[0][0]
    overall_profitable_hits = count_profitable_hits(records)
    recent_profitable_hits = count_profitable_hits(records, recent_window)
    return_concentration = positive_return_concentration(records)
    score = (
        0.40 * validation_roi_score(overall_roi)
        + 0.35 * validation_roi_score(recent_roi)
        + 0.25 * pass_ratio
    )
    failed_checks: list[str] = []
    if overall_roi <= 0.0:
        failed_checks.append("overall_roi")
    if recent_roi < 0.0:
        failed_checks.append("recent_roi")
    if median_roi < 0.0:
        failed_checks.append("median_roi")
    if pass_ratio < NESTED_VALIDATION_MIN_PASS_RATIO:
        failed_checks.append("pass_ratio")
    if overall_profitable_hits < 2:
        failed_checks.append("overall_hits")
    if recent_profitable_hits < 1:
        failed_checks.append("recent_hits")
    if return_concentration > 0.60:
        failed_checks.append("concentration")
    passed = not failed_checks
    reason = "passed" if passed else "|".join(failed_checks)
    return NestedValidationSummary(
        passed=passed,
        score=score,
        observations=observations,
        overall_roi=overall_roi,
        recent_roi=recent_roi,
        pass_ratio=pass_ratio,
        overall_profitable_hits=overall_profitable_hits,
        recent_profitable_hits=recent_profitable_hits,
        positive_return_concentration=return_concentration,
        window_rois=tuple(available_window_rois),
        reason=reason,
    )


def summarize_hit_validation(records: Sequence[ValidationRecord], budget: int) -> HitValidationSummary:
    observations = len(records)
    if observations < NESTED_VALIDATION_MIN_OBSERVATIONS:
        return HitValidationSummary(
            passed=False,
            score=0.0,
            observations=observations,
            overall_hit_score=0.0,
            recent_hit_score=0.0,
            overall_any_hit_rate=0.0,
            recent_any_hit_rate=0.0,
            overall_value_hit_rate=0.0,
            recent_value_hit_rate=0.0,
            pass_ratio=0.0,
            window_hit_scores=(),
            reason="insufficient_history",
        )
    available_window_scores: list[tuple[int, float]] = []
    for window in NESTED_VALIDATION_WINDOWS:
        if observations < max(window, NESTED_VALIDATION_MIN_OBSERVATIONS):
            continue
        available_window_scores.append((window, compute_window_hit_score(records, window)))
    if len(available_window_scores) < NESTED_VALIDATION_MIN_AVAILABLE_WINDOWS:
        return HitValidationSummary(
            passed=False,
            score=0.0,
            observations=observations,
            overall_hit_score=compute_window_hit_score(records),
            recent_hit_score=0.0,
            overall_any_hit_rate=compute_window_any_hit_rate(records),
            recent_any_hit_rate=0.0,
            overall_value_hit_rate=compute_window_value_hit_rate(records),
            recent_value_hit_rate=0.0,
            pass_ratio=0.0,
            window_hit_scores=tuple(available_window_scores),
            reason="insufficient_windows",
        )
    overall_threshold, recent_threshold, any_threshold, recent_any_threshold = hit_validation_thresholds(budget)
    overall_hit_score = compute_window_hit_score(records)
    recent_hit_score = available_window_scores[0][1]
    overall_any_hit_rate = compute_window_any_hit_rate(records)
    recent_any_hit_rate = compute_window_any_hit_rate(records, available_window_scores[0][0])
    overall_value_hit_rate = compute_window_value_hit_rate(records)
    recent_value_hit_rate = compute_window_value_hit_rate(records, available_window_scores[0][0])
    overall_strong_hit_rate = compute_window_strong_hit_rate(records)
    positive_windows = sum(window_score >= overall_threshold for _, window_score in available_window_scores)
    pass_ratio = positive_windows / len(available_window_scores)
    score = (
        0.32 * min(overall_hit_score / max(overall_threshold, 1e-6), 1.25)
        + 0.26 * min(recent_hit_score / max(recent_threshold, 1e-6), 1.25)
        + 0.18 * min(overall_any_hit_rate / max(any_threshold, 1e-6), 1.25)
        + 0.10 * min(recent_any_hit_rate / max(recent_any_threshold, 1e-6), 1.25)
        + 0.14 * min(overall_value_hit_rate / max(MIN_VALUE_HIT_RATE_OVERALL, 1e-6), 1.25)
    )
    failed_checks: list[str] = []
    if overall_hit_score < overall_threshold:
        failed_checks.append("overall_hit_score")
    if recent_hit_score < recent_threshold:
        failed_checks.append("recent_hit_score")
    if overall_any_hit_rate < any_threshold:
        failed_checks.append("overall_any_hit_rate")
    if recent_any_hit_rate < recent_any_threshold:
        failed_checks.append("recent_any_hit_rate")
    # ── 高回报门槛：拒绝低回报方案 ─────────────────────────────────────────
    if overall_value_hit_rate < MIN_VALUE_HIT_RATE_OVERALL:
        failed_checks.append("overall_value_hit_rate")
    if recent_value_hit_rate < MIN_VALUE_HIT_RATE_RECENT:
        failed_checks.append("recent_value_hit_rate")
    if overall_strong_hit_rate < MIN_STRONG_HIT_RATE_OVERALL:
        # 从未命中四等奖（200元）以上，视为永久低回报
        failed_checks.append("no_strong_hit")
    # ───────────────────────────────────────────────────────────────────────
    if pass_ratio < 0.5:
        failed_checks.append("hit_pass_ratio")
    passed = not failed_checks
    reason = "passed" if passed else "|".join(failed_checks)
    return HitValidationSummary(
        passed=passed,
        score=min(score, 1.0),
        observations=observations,
        overall_hit_score=overall_hit_score,
        recent_hit_score=recent_hit_score,
        overall_any_hit_rate=overall_any_hit_rate,
        recent_any_hit_rate=recent_any_hit_rate,
        overall_value_hit_rate=overall_value_hit_rate,
        recent_value_hit_rate=recent_value_hit_rate,
        pass_ratio=pass_ratio,
        window_hit_scores=tuple(available_window_scores),
        reason=reason,
        overall_strong_hit_rate=overall_strong_hit_rate,
    )


def combine_validation_summaries(
    roi_summary: NestedValidationSummary,
    hit_summary: HitValidationSummary,
) -> ProfileValidationSummary:
    if roi_summary.passed and hit_summary.passed:
        mode = "hybrid" if hit_summary.score >= roi_summary.score * 0.9 else "roi"
        score = max(roi_summary.score, hit_summary.score)
        return ProfileValidationSummary(
            passed=True,
            score=score,
            mode=mode,
            reason="roi+hit_passed",
            roi=roi_summary,
            hit=hit_summary,
        )
    if roi_summary.passed:
        return ProfileValidationSummary(
            passed=True,
            score=roi_summary.score,
            mode="roi",
            reason=roi_summary.reason,
            roi=roi_summary,
            hit=hit_summary,
        )
    if hit_summary.passed:
        return ProfileValidationSummary(
            passed=True,
            score=hit_summary.score,
            mode="hit",
            reason=hit_summary.reason,
            roi=roi_summary,
            hit=hit_summary,
        )
    return ProfileValidationSummary(
        passed=False,
        score=max(roi_summary.score, hit_summary.score),
        mode="none",
        reason=f"roi:{roi_summary.reason};hit:{hit_summary.reason}",
        roi=roi_summary,
        hit=hit_summary,
    )


def summarize_profile_validation(records: Sequence[ValidationRecord], budget: int) -> ProfileValidationSummary:
    roi_summary = summarize_nested_validation(records)
    hit_summary = summarize_hit_validation(records, budget)
    return combine_validation_summaries(roi_summary, hit_summary)


def get_cached_profile_validation_summary(
    cache: dict[str, tuple[int, ProfileValidationSummary]],
    profile_key: str,
    budget: int,
    records_bucket: defaultdict[str, list[ValidationRecord]],
) -> ProfileValidationSummary:
    records = records_bucket[profile_key]
    record_count = len(records)
    cached = cache.get(profile_key)
    if cached is not None and cached[0] == record_count:
        return cached[1]
    summary = summarize_profile_validation(records, budget)
    cache[profile_key] = (record_count, summary)
    return summary


def annotate_scheme_nested_validation(scheme: Scheme, summary: ProfileValidationSummary) -> None:
    scheme.metadata["nested_validation_pass"] = summary.passed
    scheme.metadata["nested_validation_mode"] = summary.mode
    scheme.metadata["nested_validation_score"] = round(summary.score, 6)
    scheme.metadata["nested_validation_reason"] = summary.reason
    scheme.metadata["nested_validation_observations"] = summary.roi.observations
    scheme.metadata["nested_validation_overall_roi"] = round(summary.roi.overall_roi, 6)
    scheme.metadata["nested_validation_recent_roi"] = round(summary.roi.recent_roi, 6)
    scheme.metadata["nested_validation_pass_ratio"] = round(summary.roi.pass_ratio, 6)
    scheme.metadata["nested_validation_overall_profitable_hits"] = summary.roi.overall_profitable_hits
    scheme.metadata["nested_validation_recent_profitable_hits"] = summary.roi.recent_profitable_hits
    scheme.metadata["nested_validation_return_concentration"] = round(summary.roi.positive_return_concentration, 6)
    scheme.metadata["nested_validation_windows"] = [
        {"window": window, "roi": round(window_roi, 6)}
        for window, window_roi in summary.roi.window_rois
    ]
    scheme.metadata["hit_validation_pass"] = summary.hit.passed
    scheme.metadata["hit_validation_score"] = round(summary.hit.score, 6)
    scheme.metadata["hit_validation_reason"] = summary.hit.reason
    scheme.metadata["hit_validation_overall_score"] = round(summary.hit.overall_hit_score, 6)
    scheme.metadata["hit_validation_recent_score"] = round(summary.hit.recent_hit_score, 6)
    scheme.metadata["hit_validation_overall_any_rate"] = round(summary.hit.overall_any_hit_rate, 6)
    scheme.metadata["hit_validation_recent_any_rate"] = round(summary.hit.recent_any_hit_rate, 6)
    scheme.metadata["hit_validation_overall_value_rate"] = round(summary.hit.overall_value_hit_rate, 6)
    scheme.metadata["hit_validation_recent_value_rate"] = round(summary.hit.recent_value_hit_rate, 6)
    scheme.metadata["hit_validation_overall_strong_rate"] = round(summary.hit.overall_strong_hit_rate, 6)
    scheme.metadata["hit_validation_pass_ratio"] = round(summary.hit.pass_ratio, 6)
    scheme.metadata["hit_validation_windows"] = [
        {"window": window, "score": round(window_score, 6)}
        for window, window_score in summary.hit.window_hit_scores
    ]


POLICY_SCORE_THRESHOLDS = {
    "family_stat_score": (0.45, 0.55, 0.65, 0.75, 0.85),
    "family_proxy_score": (0.55, 0.70, 0.85),
}


def applicable_policy_ids(
    family: str,
    family_stat_score: float,
    family_proxy_score: float,
) -> list[str]:
    policies = [f"always::{family}"]
    for threshold in POLICY_SCORE_THRESHOLDS["family_stat_score"]:
        if family_stat_score >= threshold:
            policies.append(f"stat>={threshold:.2f}::{family}")
    for threshold in POLICY_SCORE_THRESHOLDS["family_proxy_score"]:
        if family_proxy_score >= threshold:
            policies.append(f"proxy>={threshold:.2f}::{family}")
    return policies


def policy_selection_score(ema: float, count: int) -> float:
    return ema + (0.22 / math.sqrt(count + 1))


def scheme_reward_signal(result: SchemeResult, scheme: Scheme) -> float:
    budget = int(scheme.metadata.get("budget", scheme.cost)) if scheme.metadata else scheme.cost
    if budget <= 0:
        return 0.0
    utilization = scheme.cost / budget
    payout_rate = min(result.total_return / budget, 3.0)
    profit_rate = result.profit / budget
    # 若命中但全部为六等奖(5元)，视为低质量命中，奖励打折
    only_low_prize = result.winning_ticket_count > 0 and result.value_hit_tickets == 0
    issue_hit = 0.30 if only_low_prize else (1.0 if result.winning_ticket_count > 0 else 0.0)
    value_hit = 1.0 if result.value_hit_tickets > 0 else 0.0
    strong_hit = 1.0 if result.strong_hit_tickets > 0 else 0.0
    ticket_hit_ratio = result.winning_ticket_count / max(len(scheme.tickets), 1)
    match_component = min(result.average_match_score / 6.5, 1.0)
    validation_mode = str(scheme.metadata.get("nested_validation_mode", "")).strip()
    if validation_mode == "hit":
        return (
            0.16 * payout_rate
            + 0.05 * profit_rate
            + 0.12 * utilization
            + 0.26 * issue_hit   # 减少纯六等奖命中的奖励
            + 0.24 * value_hit   # 更重视五等奖及以上
            + 0.10 * strong_hit  # 重视四等奖及以上
            + 0.04 * min(ticket_hit_ratio * 4.0, 1.0)
            + 0.03 * match_component
        )
    if validation_mode == "hybrid":
        return (
            0.18 * payout_rate
            + 0.06 * profit_rate
            + 0.13 * utilization
            + 0.24 * issue_hit
            + 0.22 * value_hit
            + 0.10 * strong_hit
            + 0.04 * min(ticket_hit_ratio * 4.0, 1.0)
            + 0.03 * match_component
        )
    return (
        0.22 * payout_rate
        + 0.08 * profit_rate
        + 0.14 * utilization
        + 0.18 * issue_hit
        + 0.20 * value_hit
        + 0.10 * strong_hit
        + 0.04 * min(ticket_hit_ratio * 4.0, 1.0)
        + 0.04 * match_component
    )


def realized_roi_signal(result: SchemeResult, scheme: Scheme) -> float:
    if scheme.cost <= 0:
        return 0.0
    roi = (result.total_return / scheme.cost) - 1.0
    return min(max(roi, -1.0), 3.0)


def profile_structure_bonus(
    budget: int,
    scheme: Scheme,
    ema: float,
    proxy_score: float,
    roi_score: float,
    count: int,
) -> float:
    if budget != 1000 or scheme.family != "full_fushi" or count < 16:
        return 0.0
    red_count = len(scheme.metadata.get("reds", ()))
    blue_count = len(scheme.metadata.get("blues", ()))
    if red_count != 10 or blue_count != 2:
        return 0.0
    confidence = min((count - 16) / 28.0, 1.0)
    underpriced_signal = max(ema - proxy_score, 0.0)
    return confidence * (0.16 * roi_score + 0.10 * underpriced_signal)


def profile_selection_score(
    ema: float,
    proxy_score: float,
    utilization: float,
    count: int,
    roi_score: float = 0.0,
    validation_score: float = 0.0,
    budget: int = 0,
    scheme: Scheme | None = None,
) -> float:
    learned_weight = min(count / 20.0, 1.0)
    budget_scale = min(max(budget, 0), 1000) / 1000.0
    hit_propensity = float(scheme.metadata.get("hit_propensity", 0.0)) if scheme is not None else 0.0
    score = (
        (0.22 + 0.22 * learned_weight) * ema
        + (0.44 - 0.12 * learned_weight) * proxy_score
        + (0.04 + 0.10 * budget_scale * learned_weight) * roi_score
        + (0.05 + 0.08 * budget_scale) * validation_score
        + 0.10 * utilization
        + 0.10 * hit_propensity
        + (0.12 / math.sqrt(count + 1))
    )
    if scheme is not None:
        score += profile_structure_bonus(
            budget=budget,
            scheme=scheme,
            ema=ema,
            proxy_score=proxy_score,
            roi_score=roi_score,
            count=count,
        )
    return score


def select_family_scheme(
    budget: int,
    family_candidates: Sequence[Scheme],
    profile_stat_budget: defaultdict[str, dict[str, float]],
    profile_roi_budget: defaultdict[str, dict[str, float]],
    profile_validation_summaries: dict[str, ProfileValidationSummary] | None = None,
) -> Scheme:
    if len(family_candidates) == 1:
        return family_candidates[0]
    max_profile_count = max(
        profile_stat_budget[str(scheme.metadata["profile_key"])]["count"] for scheme in family_candidates
    )
    if max_profile_count < 12:
        return max(family_candidates, key=lambda scheme: scheme.proxy_score)
    profile_ema_scores = normalized_value_map(
        {
            str(scheme.metadata["profile_key"]): profile_stat_budget[str(scheme.metadata["profile_key"])]["ema"]
            for scheme in family_candidates
        }
    )
    profile_proxy_scores = normalized_value_map(
        {str(scheme.metadata["profile_key"]): scheme.proxy_score for scheme in family_candidates}
    )
    profile_roi_scores = normalized_value_map(
        {
            str(scheme.metadata["profile_key"]): profile_roi_budget[str(scheme.metadata["profile_key"])]["ema"]
            for scheme in family_candidates
        }
    )
    profile_validation_scores = normalized_value_map(
        {
            str(scheme.metadata["profile_key"]): (
                profile_validation_summaries[str(scheme.metadata["profile_key"])].score
                if profile_validation_summaries and str(scheme.metadata["profile_key"]) in profile_validation_summaries
                else 0.0
            )
            for scheme in family_candidates
        }
    )
    return max(
        family_candidates,
        key=lambda scheme: profile_selection_score(
            ema=profile_ema_scores.get(str(scheme.metadata["profile_key"]), 0.0),
            proxy_score=profile_proxy_scores.get(str(scheme.metadata["profile_key"]), 0.0),
            utilization=float(scheme.metadata.get("budget_utilization", 0.0)),
            count=profile_stat_budget[str(scheme.metadata["profile_key"])]["count"],
            roi_score=profile_roi_scores.get(str(scheme.metadata["profile_key"]), 0.0),
            validation_score=profile_validation_scores.get(str(scheme.metadata["profile_key"]), 0.0),
            budget=budget,
            scheme=scheme,
        ),
    )


def scheme_profile_key(scheme: Scheme) -> str:
    if scheme.family == "single_pack":
        key = f"single_pack::{scheme.metadata.get('ticket_count', len(scheme.tickets))}"
    elif scheme.family == "red_fushi":
        key = f"red_fushi::{len(scheme.metadata.get('reds', ())) }+1"
    elif scheme.family == "blue_fushi":
        key = f"blue_fushi::6+{len(scheme.metadata.get('blues', ())) }"
    elif scheme.family == "full_fushi":
        key = f"full_fushi::{len(scheme.metadata.get('reds', ())) }+{len(scheme.metadata.get('blues', ())) }"
    elif scheme.family == "dantuo":
        key = (
            f"dantuo::{len(scheme.metadata.get('dan', ())) }胆"
            f"{len(scheme.metadata.get('drag', ())) }拖+蓝{len(scheme.metadata.get('blues', ())) }"
        )
    else:
        key = scheme.family
    variant = str(scheme.metadata.get("variant", "")).strip()
    if variant:
        return f"{key}:{variant}"
    return key


def finalize_scheme(scheme: Scheme, family: str, calibrator: TicketReturnCalibrator | None) -> Scheme:
    scheme.metadata["profile_key"] = scheme_profile_key(scheme)
    scheme.metadata["base_proxy_score"] = round(scheme.proxy_score, 6)
    scheme.proxy_score = calibrate_scheme_proxy(family, scheme, calibrator)
    return scheme


def add_top_scheme(bucket: dict[str, Scheme], scheme: Scheme, limit: int) -> None:
    profile_key = str(scheme.metadata["profile_key"])
    existing = bucket.get(profile_key)
    if existing is None or scheme.proxy_score > existing.proxy_score:
        bucket[profile_key] = scheme
    if len(bucket) <= limit:
        return
    ranked = sorted(
        bucket.values(),
        key=lambda item: (
            item.proxy_score,
            float(item.metadata.get("budget_utilization", 0.0)),
            item.cost,
        ),
        reverse=True,
    )[:limit]
    bucket.clear()
    for item in ranked:
        bucket[str(item.metadata["profile_key"])] = item


FAMILY_VARIANT_LIMITS = {
    "single_pack": 10,
    "red_fushi": 12,
    "blue_fushi": 14,
    "full_fushi": 14,
    "dantuo": 12,
}


def limit_family_schemes(
    family: str,
    schemes: Sequence[Scheme],
    limit: int | None = None,
) -> list[Scheme]:
    if not schemes:
        return []
    cap = limit if limit is not None else FAMILY_VARIANT_LIMITS.get(family, len(schemes))
    if len(schemes) <= cap:
        return list(schemes)
    ranked = sorted(
        schemes,
        key=lambda scheme: (
            scheme.proxy_score,
            float(scheme.metadata.get("hit_propensity", 0.0)),
            float(scheme.metadata.get("budget_utilization", 0.0)),
            scheme.cost,
        ),
        reverse=True,
    )
    return ranked[:cap]


def build_single_pack_variant(
    pool: Sequence[Ticket],
    count: int,
    max_overlap: int,
    blue_cap: int,
) -> list[Ticket]:
    selected: list[Ticket] = []
    seen: set[tuple[tuple[int, ...], int]] = set()
    blue_usage: Counter = Counter()
    for candidate in pool:
        key = (candidate.reds, candidate.blue)
        if key in seen:
            continue
        if blue_usage[candidate.blue] >= blue_cap:
            continue
        if selected:
            recent = selected[-60:]
            overlap = max((candidate.red_mask & ticket.red_mask).bit_count() for ticket in recent)
            if overlap > max_overlap:
                continue
        selected.append(candidate)
        seen.add(key)
        blue_usage[candidate.blue] += 1
        if len(selected) >= count:
            return selected
    for candidate in pool:
        key = (candidate.reds, candidate.blue)
        if key in seen:
            continue
        selected.append(candidate)
        seen.add(key)
        if len(selected) >= count:
            break
    return selected


def build_red_anchor_variants(
    ranked_reds: Sequence[int],
    red_scores: dict[int, float],
    count: int,
) -> list[tuple[str, tuple[int, ...]]]:
    variants: list[tuple[str, tuple[int, ...]]] = []
    seen: set[tuple[int, ...]] = set()
    for label, start in (("core", 0), ("shift1", 1), ("shift2", 2)):
        if len(ranked_reds) - start < count:
            continue
        pool = choose_best_red_pool(ranked_reds[start:], red_scores, count)
        if pool in seen:
            continue
        seen.add(pool)
        variants.append((label, pool))
    return variants


def interleave_tickets_by_blue(pool: Sequence[Ticket]) -> list[Ticket]:
    grouped: defaultdict[int, list[Ticket]] = defaultdict(list)
    for ticket in pool:
        grouped[ticket.blue].append(ticket)
    ordered_blues = sorted(
        grouped,
        key=lambda blue: (-grouped[blue][0].score, blue),
    )
    interleaved: list[Ticket] = []
    depth = 0
    while True:
        added = False
        for blue in ordered_blues:
            tickets = grouped[blue]
            if depth < len(tickets):
                interleaved.append(tickets[depth])
                added = True
        if not added:
            break
        depth += 1
    return interleaved


def describe_variant(variant: str) -> str:
    return {
        "core": "稳态核心",
        "low_overlap": "低重叠分散",
        "blue_spread": "蓝球分散",
        "reserve_floor": "保守仓位",
        "reserve_mid": "中性仓位",
        "entry_loose": "宽松入场",
        "entry_tight": "严格入场",
        "gate_loose": "门控宽松",
        "gate_tight": "门控严格",
        "form_hot": "历史强势",
        "form_confirmed": "历史确认",
        "hit_warm": "命中记忆",
        "hit_hot": "高回报记忆",
        "shift1": "红球偏移一档",
        "shift2": "红球偏移二档",
    }.get(variant, variant)


CONDITIONAL_ENTRY_THRESHOLDS = {
    "single_pack": (-0.62, -0.48),
    "red_fushi": (-0.56, -0.40),
    "blue_fushi": (-0.34, -0.22),
    "full_fushi": (-0.54, -0.38),
}

GATE_ENTRY_THRESHOLDS = {
    "single_pack": (-0.10, 0.02),
    "red_fushi": (-0.06, 0.05),
    "blue_fushi": (-0.08, 0.03),
    "full_fushi": (-0.08, 0.04),
    "dantuo": (-0.05, 0.06),
}


def clone_scheme_with_variant(
    scheme: Scheme,
    variant_suffix: str,
    calibrator: TicketReturnCalibrator | None,
) -> Scheme:
    metadata = dict(scheme.metadata)
    base_variant = str(metadata.get("variant", "")).strip()
    metadata["variant"] = f"{base_variant}|{variant_suffix}" if base_variant else variant_suffix
    cloned = Scheme(
        family=scheme.family,
        description=f"{scheme.description}（{describe_variant(variant_suffix)}）",
        cost=scheme.cost,
        tickets=list(scheme.tickets),
        proxy_score=float(scheme.metadata.get("base_proxy_score", scheme.proxy_score)),
        metadata=metadata,
    )
    return finalize_scheme(cloned, scheme.family, calibrator)


def expand_conditional_entry_variants(
    family: str,
    schemes: Sequence[Scheme],
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    thresholds = CONDITIONAL_ENTRY_THRESHOLDS.get(family)
    if thresholds is None or calibrator is None:
        return list(schemes)
    expanded: dict[str, Scheme] = {str(scheme.metadata["profile_key"]): scheme for scheme in schemes}
    loose_threshold, tight_threshold = thresholds
    for scheme in schemes:
        expected_profit_ratio = float(scheme.metadata.get("expected_profit_ratio", -1.0))
        if expected_profit_ratio >= loose_threshold:
            clone = clone_scheme_with_variant(scheme, "entry_loose", calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
        if expected_profit_ratio >= tight_threshold:
            clone = clone_scheme_with_variant(scheme, "entry_tight", calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
    return list(expanded.values())


def score_focus_gap(scores: dict[int, float], head: int, tail: int) -> float:
    ranked = sorted(scores, key=scores.get, reverse=True)
    head_scores = [scores[number] for number in ranked[:head]]
    tail_scores = [scores[number] for number in ranked[head:tail]]
    if not head_scores or not tail_scores:
        return 0.0
    return average(head_scores) - average(tail_scores)


def build_regime_labels(
    red_arm_weights: dict[str, float],
    blue_arm_weights: dict[str, float],
    red_source_weights: dict[str, float],
    blue_source_weights: dict[str, float],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
) -> dict[str, tuple[str, ...]]:
    red_labels: list[str] = []
    blue_labels: list[str] = []

    top_red_arm = max(red_arm_weights, key=red_arm_weights.get)
    if red_arm_weights[top_red_arm] >= 0.22:
        red_labels.append(f"redarm_{top_red_arm}")
    top_blue_arm = max(blue_arm_weights, key=blue_arm_weights.get)
    if blue_arm_weights[top_blue_arm] >= 0.28:
        blue_labels.append(f"bluearm_{top_blue_arm}")

    top_red_source = max(red_source_weights, key=red_source_weights.get)
    if red_source_weights[top_red_source] >= 0.64:
        red_labels.append(f"redsrc_{top_red_source}")
    top_blue_source = max(blue_source_weights, key=blue_source_weights.get)
    if blue_source_weights[top_blue_source] >= 0.64:
        blue_labels.append(f"bluesrc_{top_blue_source}")

    red_focus = score_focus_gap(red_scores, 6, 14)
    if red_focus >= 0.17:
        red_labels.append("redfocus_tight")
    elif red_focus <= 0.10:
        red_labels.append("redfocus_wide")

    blue_focus = score_focus_gap(blue_scores, 3, 8)
    if blue_focus >= 0.22:
        blue_labels.append("bluefocus_tight")
    elif blue_focus <= 0.12:
        blue_labels.append("bluefocus_wide")

    def unique(labels: Sequence[str]) -> tuple[str, ...]:
        seen: set[str] = set()
        ordered: list[str] = []
        for label in labels:
            if label in seen:
                continue
            seen.add(label)
            ordered.append(label)
        return tuple(ordered)

    return {
        "single_pack": unique(red_labels[:2] + blue_labels[:2])[:4],
        "red_fushi": unique(red_labels)[:3],
        "blue_fushi": unique(blue_labels)[:3],
        "full_fushi": unique(red_labels[:2] + blue_labels[:1])[:3],
        "dantuo": unique(red_labels[:1] + blue_labels[:1])[:2],
    }


def expand_regime_variants(
    family: str,
    schemes: Sequence[Scheme],
    regime_labels: dict[str, tuple[str, ...]] | None,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    if calibrator is None or not regime_labels:
        return list(schemes)
    labels = regime_labels.get(family, ())
    if not labels:
        return list(schemes)
    expanded: dict[str, Scheme] = {str(scheme.metadata["profile_key"]): scheme for scheme in schemes}
    combo_labels: list[str] = []
    focus_label = next((label for label in labels if "focus" in label), None)
    if focus_label is not None:
        partner_label = next(
            (
                label
                for label in reversed(labels)
                if label != focus_label and ("src_" in label or "arm_" in label)
            ),
            None,
        )
        if partner_label is not None:
            combo_labels.append(f"state_{partner_label}__{focus_label}")
    elif len(labels) >= 2:
        combo_labels.append(f"state_{labels[0]}__{labels[1]}")
    for scheme in schemes:
        for label in labels:
            clone = clone_scheme_with_variant(scheme, f"state_{label}", calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
        for combo_label in combo_labels:
            clone = clone_scheme_with_variant(scheme, combo_label, calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
    return list(expanded.values())


def expand_gate_variants(
    family: str,
    schemes: Sequence[Scheme],
    predicted_gate_roi: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    thresholds = GATE_ENTRY_THRESHOLDS.get(family)
    if calibrator is None or thresholds is None:
        return list(schemes)
    expanded: dict[str, Scheme] = {str(scheme.metadata["profile_key"]): scheme for scheme in schemes}
    loose_threshold, tight_threshold = thresholds
    for scheme in schemes:
        if predicted_gate_roi >= loose_threshold:
            clone = clone_scheme_with_variant(scheme, "gate_loose", calibrator)
            clone.metadata["predicted_gate_roi"] = round(predicted_gate_roi, 6)
            expanded[str(clone.metadata["profile_key"])] = clone
        if predicted_gate_roi >= tight_threshold:
            clone = clone_scheme_with_variant(scheme, "gate_tight", calibrator)
            clone.metadata["predicted_gate_roi"] = round(predicted_gate_roi, 6)
            expanded[str(clone.metadata["profile_key"])] = clone
    return list(expanded.values())


def expand_form_variants(
    schemes: Sequence[Scheme],
    profile_stat_budget: defaultdict[str, dict[str, float]],
    profile_roi_budget: defaultdict[str, dict[str, float]],
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    if calibrator is None or not schemes:
        return list(schemes)
    profile_ema_scores = normalized_value_map(
        {
            str(scheme.metadata["profile_key"]): profile_stat_budget[str(scheme.metadata["profile_key"])]["ema"]
            for scheme in schemes
        }
    )
    profile_roi_scores = normalized_value_map(
        {
            str(scheme.metadata["profile_key"]): profile_roi_budget[str(scheme.metadata["profile_key"])]["ema"]
            for scheme in schemes
        }
    )
    expanded: dict[str, Scheme] = {str(scheme.metadata["profile_key"]): scheme for scheme in schemes}
    for scheme in schemes:
        profile_key = str(scheme.metadata["profile_key"])
        count = int(profile_stat_budget[profile_key]["count"])
        if count < 16:
            continue
        ema_score = profile_ema_scores.get(profile_key, 0.0)
        roi_score = profile_roi_scores.get(profile_key, 0.0)
        if ema_score >= 0.78:
            clone = clone_scheme_with_variant(scheme, "form_hot", calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
        if ema_score >= 0.66 and roi_score >= 0.58:
            clone = clone_scheme_with_variant(scheme, "form_confirmed", calibrator)
            expanded[str(clone.metadata["profile_key"])] = clone
    return list(expanded.values())


def compute_reward_hit_alignment(
    current_context: np.ndarray,
    snapshots: Sequence[RewardHitSnapshot],
    top_k: int = 12,
) -> float:
    if not snapshots:
        return 0.0
    total_snapshots = len(snapshots)
    weighted_scores: list[float] = []
    for index, snapshot in enumerate(snapshots):
        # 结合欧氏+余弦相似度（与之前build_similarity_feature_map一致）
        euclidean_dist = float(np.linalg.norm(current_context - snapshot.context_vector))
        cos_sim = cosine_similarity(current_context, snapshot.context_vector)
        euclidean_sim = math.exp(-3.2 * euclidean_dist)
        cosine_component = (cos_sim + 1.0) / 2.0
        context_sim = 0.60 * euclidean_sim + 0.40 * cosine_component
        recency_weight = 0.986 ** (total_snapshots - index - 1)
        # 大幅增加高回报历史期的权重，使alignment对真正的高回报更敏感
        reward_weight = 0.40 + 0.50 * min(max(snapshot.reward_ratio, 0.0), 3.0)
        profit_weight = 1.0 + 0.35 * max(snapshot.profit_ratio, 0.0)  # 旧:0.20
        similarity = context_sim * recency_weight * reward_weight * profit_weight
        weighted_scores.append(similarity)
    top_scores = sorted(weighted_scores, reverse=True)[:top_k]
    return min(sum(top_scores) / 1.20, 1.0)


def expand_hit_variants(
    schemes: Sequence[Scheme],
    hit_alignment: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    if calibrator is None or hit_alignment < 0.08:
        return list(schemes)
    expanded: dict[str, Scheme] = {str(scheme.metadata["profile_key"]): scheme for scheme in schemes}
    for scheme in schemes:
        if hit_alignment >= 0.14:   # 提高门槛(旧:0.08)
            clone = clone_scheme_with_variant(scheme, "hit_warm", calibrator)
            clone.metadata["history_hit_alignment"] = round(hit_alignment, 6)
            expanded[str(clone.metadata["profile_key"])] = clone
        if hit_alignment >= 0.24:   # 提高门槛(旧:0.16)
            clone = clone_scheme_with_variant(scheme, "hit_hot", calibrator)
            clone.metadata["history_hit_alignment"] = round(hit_alignment, 6)
            expanded[str(clone.metadata["profile_key"])] = clone
    return list(expanded.values())


def family_cooldown_length(budget: int) -> int:
    # 更长冷却期：连续亏损后需要更多周期休整
    if budget <= 50:
        return 18   # 旧:10
    if budget <= 200:
        return 14   # 旧:8
    return 10       # 旧:6


def family_hit_reentry_threshold(budget: int) -> float:
    # 更严格再入场门槛：冷却中必须有很高命中率才能提前退出冷却
    if budget <= 50:
        return 0.94  # 旧:0.88
    if budget <= 200:
        return 0.75  # 旧:0.58
    if budget <= 500:
        return 0.60  # 旧:0.42
    return 0.50      # 旧:0.32


def family_is_cooldown_blocked(
    state: FamilyLiveState,
    budget: int,
    family_candidates: Sequence[Scheme],
    profile_validation_summaries: dict[str, ProfileValidationSummary],
) -> bool:
    if state.cooldown <= 0:
        return False
    if not family_candidates:
        return True
    strongest_hit_summary: ProfileValidationSummary | None = None
    for scheme in family_candidates:
        profile_key = str(scheme.metadata["profile_key"])
        summary = profile_validation_summaries.get(profile_key)
        if summary is None or not summary.passed or summary.mode not in {"hit", "hybrid"}:
            continue
        if strongest_hit_summary is None or (
            summary.hit.recent_hit_score,
            summary.hit.recent_any_hit_rate,
            summary.hit.score,
        ) > (
            strongest_hit_summary.hit.recent_hit_score,
            strongest_hit_summary.hit.recent_any_hit_rate,
            strongest_hit_summary.hit.score,
        ):
            strongest_hit_summary = summary
    if strongest_hit_summary is None:
        return True
    if (
        state.hitless_streak <= 1
        and state.ema_hit >= family_hit_reentry_threshold(budget)
        and strongest_hit_summary.hit.recent_any_hit_rate >= family_hit_reentry_threshold(budget)
    ):
        return False
    if (
        state.hitless_streak == 0
        and state.ema_value_hit >= 0.03
        and strongest_hit_summary.hit.recent_value_hit_rate >= 0.03
    ):
        return False
    return True


def update_family_live_state(
    state: FamilyLiveState,
    budget: int,
    scheme: Scheme,
    result: SchemeResult,
) -> None:
    if scheme.cost <= 0:
        return
    profit_ratio = result.profit / scheme.cost
    return_ratio = result.total_return / scheme.cost
    any_hit = 1.0 if result.winning_ticket_count > 0 else 0.0
    value_hit = 1.0 if result.value_hit_tickets > 0 else 0.0
    validation_mode = str(scheme.metadata.get("nested_validation_mode", "")).strip()
    hit_oriented = validation_mode in {"hit", "hybrid"} or bool(scheme.metadata.get("hit_validation_pass"))
    state.ema_profit = profit_ratio if state.count == 0 else 0.65 * state.ema_profit + 0.35 * profit_ratio
    state.ema_return = return_ratio if state.count == 0 else 0.65 * state.ema_return + 0.35 * return_ratio
    state.ema_hit = any_hit if state.count == 0 else 0.72 * state.ema_hit + 0.28 * any_hit
    state.ema_value_hit = value_hit if state.count == 0 else 0.78 * state.ema_value_hit + 0.22 * value_hit
    state.count += 1
    state.loss_streak = state.loss_streak + 1 if result.profit < 0 else 0
    state.hitless_streak = state.hitless_streak + 1 if any_hit <= 0.0 else 0
    if hit_oriented:
        if (
            # 命中信号偏弱时更快进入冷却
            (state.count >= 4 and state.ema_hit <= 0.18 and state.ema_value_hit <= 0.03 and state.ema_return <= 0.15)
            or (state.hitless_streak >= 2 and state.ema_hit <= 0.32 and state.ema_return <= 0.20)
            or (state.count >= 8 and state.ema_value_hit <= 0.015)  # 长期只中六等奖
        ):
            state.cooldown = max(state.cooldown, family_cooldown_length(budget))  # 全额冷却(旧:半额)
        return
    if (
        (state.count >= 1 and state.ema_profit <= -0.65 and state.ema_return <= 0.35)
        or (state.loss_streak >= 1 and state.ema_profit <= -0.55)  # 旧:亏损2期才冷却
        or (state.count >= 3 and state.ema_return <= 0.20)         # 新：长期低回报即冷却
    ):
        state.cooldown = max(state.cooldown, family_cooldown_length(budget))


def choose_best_red_pool(
    ranked_reds: Sequence[int],
    red_scores: dict[int, float],
    count: int,
) -> tuple[int, ...]:
    universe = tuple(ranked_reds[:14])
    best_combo: tuple[int, ...] | None = None
    best_score = -10.0
    for combo in combinations(universe, count):
        combo = tuple(sorted(combo))
        odd_count = sum(number % 2 for number in combo)
        zones = zone_split(combo)
        spread = combo[-1] - combo[0]
        consecutive = count_consecutive_pairs(combo)
        score = sum(red_scores[number] for number in combo) / count
        score -= 0.030 * abs(odd_count - count / 2)
        score += 0.020 * len([zone for zone in zones if zone > 0])
        score -= 0.020 * max(zones)
        score -= 0.015 * consecutive
        score += 0.015 * min(spread / 32.0, 1.0)
        if score > best_score:
            best_score = score
            best_combo = combo
    if best_combo is None:
        return tuple(sorted(ranked_reds[:count]))
    return best_combo


def build_red_pool_cache(
    ranked_reds: Sequence[int],
    red_scores: dict[int, float],
    counts: Iterable[int],
) -> dict[int, tuple[int, ...]]:
    return {
        count: choose_best_red_pool(ranked_reds, red_scores, count)
        for count in sorted(set(counts))
    }


def build_candidate_single_tickets(
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    max_pool_size: int,
) -> list[Ticket]:
    ranked_reds = sorted(red_scores, key=lambda number: (-red_scores[number], number))
    ranked_blues = sorted(blue_scores, key=lambda number: (-blue_scores[number], number))
    red_universe = ranked_reds[:16]   # 扩大宇宙(旧:13←11)，2026032期红01排名15未覆盖
    blue_universe = ranked_blues[:7]  # 扩大蓝球宇宙(旧:6←5)
    red_cache: dict[tuple[int, ...], float] = {}
    candidates: list[Ticket] = []
    for reds in combinations(red_universe, 6):
        red_tuple = tuple(sorted(reds))
        for blue in blue_universe:
            score = ticket_score(
                red_tuple,
                blue,
                red_scores,
                blue_scores,
                shape_context,
                pair_scores,
                max_pair_score,
                red_cache,
            )
            candidates.append(Ticket(reds=red_tuple, blue=blue, score=score))
    candidates.sort(key=lambda ticket: (-ticket.score, ticket.reds, ticket.blue))
    selected: list[Ticket] = []
    seen: set[tuple[tuple[int, ...], int]] = set()
    # 逐步放宽重叠限制：先只选低重叠的高质量票，再逐步填充
    overlap_limits = [2, 3, 4, 5, 6]
    for overlap_limit in overlap_limits:
        for candidate in candidates:
            key = (candidate.reds, candidate.blue)
            if key in seen:
                continue
            if not selected:
                selected.append(candidate)
                seen.add(key)
                if len(selected) >= max_pool_size:
                    return selected
                continue
            # 修复：重叠检查必须在内层循环内执行
            recent = selected[-50:]
            max_overlap = max(
                (candidate.red_mask & ticket.red_mask).bit_count()
                for ticket in recent
            )
            if max_overlap > overlap_limit:
                continue
            seen.add(key)
            selected.append(candidate)
            if len(selected) >= max_pool_size:
                return selected
    # 若仍未填满，不限重叠地追加
    if len(selected) < max_pool_size:
        for candidate in candidates:
            key = (candidate.reds, candidate.blue)
            if key in seen:
                continue
            selected.append(candidate)
            seen.add(key)
            if len(selected) >= max_pool_size:
                break
    return selected


def estimate_scheme_base_score(tickets: Sequence[Ticket], budget: int, cost: int) -> float:
    if not tickets:
        return -1.0
    ticket_scores = [ticket.score for ticket in tickets]
    average_score = sum(ticket_scores) / len(ticket_scores)
    high_score = max(ticket_scores)
    blue_coverage = len({ticket.blue for ticket in tickets}) / max(1, min(len(tickets), BLUE_MAX))
    red_coverage = len({number for ticket in tickets for number in ticket.reds}) / RED_MAX
    ticket_density = len(tickets) / max(1, budget // TICKET_COST)
    return (
        average_score * (0.70 + 0.10 * (cost / budget))
        + 0.06 * high_score
        + 0.10 * blue_coverage
        + 0.08 * red_coverage
        + 0.06 * ticket_density
    )


def scheme_hit_propensity(scheme: Scheme) -> float:
    ticket_count = len(scheme.tickets)
    if ticket_count <= 0:
        return 0.0
    budget = int(scheme.metadata.get("budget", scheme.cost)) if scheme.metadata else scheme.cost
    utilization = scheme.cost / budget if budget else 0.0
    blue_coverage = len({ticket.blue for ticket in scheme.tickets}) / BLUE_MAX
    red_coverage = len({number for ticket in scheme.tickets for number in ticket.reds}) / RED_MAX
    ticket_density = ticket_count / max(1, budget // TICKET_COST)
    return min(
        0.38 * blue_coverage
        + 0.24 * red_coverage
        + 0.18 * ticket_density
        + 0.20 * utilization,
        1.0,
    )


def calibrate_scheme_proxy(
    family: str,
    scheme: Scheme,
    calibrator: TicketReturnCalibrator | None,
) -> float:
    base_score = scheme.metadata.get("base_proxy_score", scheme.proxy_score)
    budget = int(scheme.metadata.get("budget", scheme.cost)) if scheme.metadata else scheme.cost
    utilization = scheme.cost / budget if budget else 1.0
    if calibrator is None:
        scheme.metadata["expected_return"] = 0.0
        scheme.metadata["expected_profit_ratio"] = -1.0
        scheme.metadata["expected_return_ratio"] = 0.0
        scheme.metadata["budget_utilization"] = round(utilization, 6)
        return base_score
    expected_return = calibrator.predict_scheme_return(family, scheme)
    expected_profit_ratio = (expected_return - scheme.cost) / scheme.cost if scheme.cost else -1.0
    expected_return_ratio = expected_return / budget if budget else 0.0
    ticket_scores = [ticket.score for ticket in scheme.tickets]
    average_score = sum(ticket_scores) / len(ticket_scores) if ticket_scores else 0.0
    max_score = max(ticket_scores) if ticket_scores else 0.0
    scheme.metadata["expected_return"] = round(expected_return, 4)
    scheme.metadata["expected_profit_ratio"] = round(expected_profit_ratio, 6)
    scheme.metadata["expected_return_ratio"] = round(expected_return_ratio, 6)
    scheme.metadata["budget_utilization"] = round(utilization, 6)
    hit_propensity = scheme_hit_propensity(scheme)
    scheme.metadata["hit_propensity"] = round(hit_propensity, 6)
    return (
        0.24 * expected_return_ratio
        + 0.08 * expected_profit_ratio
        + 0.16 * utilization
        + 0.18 * average_score
        + 0.10 * max_score
        + 0.24 * hit_propensity
    )


def make_single_pack_schemes(
    pool: Sequence[Ticket],
    budget: int,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    count = budget // TICKET_COST
    if count <= 0:
        return []
    variants: list[tuple[str, list[Ticket]]] = []
    primary = list(pool[:count])
    if primary:
        variants.append(("core", primary))
    low_overlap_cap = max(2, min(5, 3 + count // 80))
    blue_cap = max(1, math.ceil(count / 10))
    low_overlap = build_single_pack_variant(pool, count, max_overlap=low_overlap_cap, blue_cap=blue_cap)
    if low_overlap:
        variants.append(("low_overlap", low_overlap))
    blue_spread_pool = interleave_tickets_by_blue(pool)
    blue_spread = build_single_pack_variant(
        blue_spread_pool,
        count,
        max_overlap=max(low_overlap_cap, 4),
        blue_cap=max(1, math.ceil(count / 14)),
    )
    if blue_spread:
        variants.append(("blue_spread", blue_spread))
    floor_utilization = budget_utilization_floor(budget)
    reserve_floor_count = min(count, math.ceil((budget * floor_utilization) / TICKET_COST))
    if 0 < reserve_floor_count < count:
        reserve_floor = build_single_pack_variant(
            pool,
            reserve_floor_count,
            max_overlap=max(2, low_overlap_cap - 1),
            blue_cap=max(1, math.ceil(reserve_floor_count / 12)),
        )
        if reserve_floor:
            variants.append(("reserve_floor", reserve_floor))
    reserve_mid_utilization = min(0.86, max(floor_utilization + 0.12, 0.78))
    reserve_mid_count = min(count, math.ceil((budget * reserve_mid_utilization) / TICKET_COST))
    if 0 < reserve_mid_count < count and reserve_mid_count != reserve_floor_count:
        reserve_mid = build_single_pack_variant(
            blue_spread_pool,
            reserve_mid_count,
            max_overlap=max(3, low_overlap_cap),
            blue_cap=max(1, math.ceil(reserve_mid_count / 14)),
        )
        if reserve_mid:
            variants.append(("reserve_mid", reserve_mid))
    schemes: dict[str, Scheme] = {}
    seen_ticket_sets: set[tuple[tuple[tuple[int, ...], int], ...]] = set()
    for variant, tickets in variants:
        ticket_signature = tuple((ticket.reds, ticket.blue) for ticket in tickets)
        if ticket_signature in seen_ticket_sets:
            continue
        seen_ticket_sets.add(ticket_signature)
        cost = len(tickets) * TICKET_COST
        scheme = Scheme(
            family="single_pack",
            description=f"单式 {len(tickets)} 注（{describe_variant(variant)}）",
            cost=cost,
            tickets=tickets,
            proxy_score=estimate_scheme_base_score(tickets, budget, cost),
            metadata={"ticket_count": len(tickets), "budget": budget, "variant": variant},
        )
        add_top_scheme(schemes, finalize_scheme(scheme, "single_pack", calibrator), limit=4)
    return list(schemes.values())


def make_red_fushi_schemes(
    budget: int,
    ranked_reds: Sequence[int],
    red_pool_cache: dict[int, tuple[int, ...]],
    ranked_blues: Sequence[int],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    red_cache: dict[tuple[int, ...], float] = {}
    top_schemes: dict[str, Scheme] = {}
    for red_count in range(7, 13):
        ticket_count = math.comb(red_count, 6)
        cost = ticket_count * TICKET_COST
        if cost > budget:
            continue
        blue = ranked_blues[0]
        variants = build_red_anchor_variants(ranked_reds, red_scores, red_count)
        if not variants:
            variants = [("core", red_pool_cache[red_count])]
        for variant, red_pool in variants:
            tickets = [
                Ticket(
                    reds=combo,
                    blue=blue,
                    score=ticket_score(
                        combo,
                        blue,
                        red_scores,
                        blue_scores,
                        shape_context,
                        pair_scores,
                        max_pair_score,
                        red_cache,
                    ),
                )
                for combo in combinations(red_pool, 6)
            ]
            scheme = Scheme(
                family="red_fushi",
                description=f"红球复式 {red_count}+1（{describe_variant(variant)}）",
                cost=cost,
                tickets=tickets,
                proxy_score=estimate_scheme_base_score(tickets, budget, cost),
                metadata={"reds": red_pool, "blue": blue, "budget": budget, "variant": variant},
            )
            add_top_scheme(top_schemes, finalize_scheme(scheme, "red_fushi", calibrator), limit=12)
    return list(top_schemes.values())


def make_blue_fushi_schemes(
    budget: int,
    ranked_reds: Sequence[int],
    red_pool_cache: dict[int, tuple[int, ...]],
    ranked_blues: Sequence[int],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    red_cache: dict[tuple[int, ...], float] = {}
    top_schemes: dict[str, Scheme] = {}
    red_variants = build_red_anchor_variants(ranked_reds, red_scores, 6)
    if not red_variants:
        red_variants = [("core", red_pool_cache[6])]
    # 限制蓝球数量上限，拒绝宽蓝覆盖（超过一半蓝球池即视为低效撒网）
    blue_count_candidates = range(2, min(BLUE_FUSHI_MAX_BLUES, BLUE_MAX) + 1)
    for blue_count in blue_count_candidates:
        cost = blue_count * TICKET_COST
        if cost > budget:
            continue
        blue_pool = tuple(sorted(ranked_blues[:blue_count]))
        for variant, red_pool in red_variants:
            tickets = [
                Ticket(
                    reds=red_pool,
                    blue=blue,
                    score=ticket_score(
                        red_pool,
                        blue,
                        red_scores,
                        blue_scores,
                        shape_context,
                        pair_scores,
                        max_pair_score,
                        red_cache,
                    ),
                )
                for blue in blue_pool
            ]
            scheme = Scheme(
                family="blue_fushi",
                description=f"蓝球复式 6+{blue_count}（{describe_variant(variant)}）",
                cost=cost,
                tickets=tickets,
                proxy_score=estimate_scheme_base_score(tickets, budget, cost),
                metadata={"reds": red_pool, "blues": blue_pool, "budget": budget, "variant": variant},
            )
            add_top_scheme(top_schemes, finalize_scheme(scheme, "blue_fushi", calibrator), limit=24)
    return list(top_schemes.values())


def make_full_fushi_schemes(
    budget: int,
    ranked_reds: Sequence[int],
    red_pool_cache: dict[int, tuple[int, ...]],
    ranked_blues: Sequence[int],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    red_cache: dict[tuple[int, ...], float] = {}
    top_schemes: dict[str, Scheme] = {}
    for red_count in range(7, 11):
        red_ticket_count = math.comb(red_count, 6)
        red_variants = build_red_anchor_variants(ranked_reds, red_scores, red_count) if red_count <= 8 else []
        if not red_variants:
            red_variants = [("core", red_pool_cache[red_count])]
        for blue_count in range(2, min(8, BLUE_MAX) + 1):
            cost = red_ticket_count * blue_count * TICKET_COST
            if cost > budget:
                continue
            blue_pool = tuple(sorted(ranked_blues[:blue_count]))
            for variant, red_pool in red_variants:
                tickets = []
                for combo in combinations(red_pool, 6):
                    for blue in blue_pool:
                        tickets.append(
                            Ticket(
                                reds=combo,
                                blue=blue,
                                score=ticket_score(
                                    combo,
                                    blue,
                                    red_scores,
                                    blue_scores,
                                    shape_context,
                                    pair_scores,
                                    max_pair_score,
                                    red_cache,
                                ),
                            )
                        )
                scheme = Scheme(
                    family="full_fushi",
                    description=f"全复式 {red_count}+{blue_count}（{describe_variant(variant)}）",
                    cost=cost,
                    tickets=tickets,
                    proxy_score=estimate_scheme_base_score(tickets, budget, cost),
                    metadata={"reds": red_pool, "blues": blue_pool, "budget": budget, "variant": variant},
                )
                add_top_scheme(top_schemes, finalize_scheme(scheme, "full_fushi", calibrator), limit=28)
    return list(top_schemes.values())


def make_dantuo_schemes(
    budget: int,
    red_pool_cache: dict[int, tuple[int, ...]],
    ranked_blues: Sequence[int],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    calibrator: TicketReturnCalibrator | None,
) -> list[Scheme]:
    red_cache: dict[tuple[int, ...], float] = {}
    top_schemes: dict[str, Scheme] = {}
    for dan_count in range(2, 6):
        for total_red_count in range(dan_count + 2, 13):
            drag_count = total_red_count - dan_count
            need_from_drag = 6 - dan_count
            if drag_count < need_from_drag + 1:
                continue
            red_pool = red_pool_cache[total_red_count]
            sorted_by_score = sorted(red_pool, key=lambda number: (-red_scores[number], number))
            dan = tuple(sorted(sorted_by_score[:dan_count]))
            drag = tuple(sorted(number for number in red_pool if number not in dan))
            combo_count = math.comb(len(drag), need_from_drag)
            for blue_count in range(1, min(12, BLUE_MAX) + 1):
                cost = combo_count * blue_count * TICKET_COST
                if cost > budget:
                    continue
                blue_pool = tuple(sorted(ranked_blues[:blue_count]))
                tickets = []
                for drag_combo in combinations(drag, need_from_drag):
                    reds = tuple(sorted(dan + drag_combo))
                    for blue in blue_pool:
                        tickets.append(
                            Ticket(
                                reds=reds,
                                blue=blue,
                                score=ticket_score(
                                    reds,
                                    blue,
                                    red_scores,
                                    blue_scores,
                                    shape_context,
                                    pair_scores,
                                    max_pair_score,
                                    red_cache,
                                ),
                            )
                        )
                scheme = Scheme(
                    family="dantuo",
                    description=f"胆拖 {dan_count}胆{len(drag)}拖 + 蓝{blue_count}",
                    cost=cost,
                    tickets=tickets,
                    proxy_score=estimate_scheme_base_score(tickets, budget, cost),
                    metadata={"dan": dan, "drag": drag, "blues": blue_pool, "budget": budget},
                )
                add_top_scheme(top_schemes, finalize_scheme(scheme, "dantuo", calibrator), limit=3)
    return list(top_schemes.values())


def build_budget_schemes(
    budget: int,
    ranked_reds: Sequence[int],
    ranked_blues: Sequence[int],
    red_pool_cache: dict[int, tuple[int, ...]],
    red_scores: dict[int, float],
    blue_scores: dict[int, float],
    shape_context: dict[str, object],
    pair_scores: dict[tuple[int, int], float],
    max_pair_score: float,
    single_pool: Sequence[Ticket],
    regime_labels: dict[str, tuple[str, ...]] | None,
    calibrator: TicketReturnCalibrator | None,
) -> dict[str, list[Scheme]]:
    family_candidates = {
        "single_pack": expand_regime_variants(
            "single_pack",
            expand_conditional_entry_variants(
                "single_pack",
                make_single_pack_schemes(single_pool, budget, calibrator),
                calibrator,
            ),
            regime_labels,
            calibrator,
        ),
        "red_fushi": expand_regime_variants(
            "red_fushi",
            make_red_fushi_schemes(
                budget,
                ranked_reds,
                red_pool_cache,
                ranked_blues,
                red_scores,
                blue_scores,
                shape_context,
                pair_scores,
                max_pair_score,
                calibrator,
            ),
            regime_labels,
            calibrator,
        ),
        "blue_fushi": expand_regime_variants(
            "blue_fushi",
            expand_conditional_entry_variants(
                "blue_fushi",
                make_blue_fushi_schemes(
                    budget,
                    ranked_reds,
                    red_pool_cache,
                    ranked_blues,
                    red_scores,
                    blue_scores,
                    shape_context,
                    pair_scores,
                    max_pair_score,
                    calibrator,
                ),
                calibrator,
            ),
            regime_labels,
            calibrator,
        ),
        "full_fushi": expand_regime_variants(
            "full_fushi",
            expand_conditional_entry_variants(
                "full_fushi",
                make_full_fushi_schemes(
                    budget,
                    ranked_reds,
                    red_pool_cache,
                    ranked_blues,
                    red_scores,
                    blue_scores,
                    shape_context,
                    pair_scores,
                    max_pair_score,
                    calibrator,
                ),
                calibrator,
            ),
            regime_labels,
            calibrator,
        ),
        "dantuo": expand_regime_variants(
            "dantuo",
            make_dantuo_schemes(
                budget,
                red_pool_cache,
                ranked_blues,
                red_scores,
                blue_scores,
                shape_context,
                pair_scores,
                max_pair_score,
                calibrator,
            ),
            regime_labels,
            calibrator,
        ),
    }
    family_candidates["red_fushi"] = expand_conditional_entry_variants(
        "red_fushi",
        family_candidates["red_fushi"],
        calibrator,
    )
    schemes = {
        family: limit_family_schemes(
            family,
            [scheme for scheme in candidates if scheme is not None],
        )
        for family, candidates in family_candidates.items()
        if candidates
    }
    utilization_floor = budget_utilization_floor(budget)
    eligible = {
        family: family_schemes
        for family, scheme in schemes.items()
        if (family_schemes := [item for item in scheme if budget <= 0 or (item.cost / budget) >= utilization_floor])
    }
    return eligible or schemes


PRIZE_TYPE_BY_NAME = {
    "一等奖": 1,
    "二等奖": 2,
    "三等奖": 3,
    "四等奖": 4,
    "五等奖": 5,
    "六等奖": 6,
}


def prize_payout(draw: Draw, prize_name: str) -> int:
    if prize_name in FIXED_PAYOUTS:
        return FIXED_PAYOUTS[prize_name]
    if draw.official is None:
        return 0
    prize_type = PRIZE_TYPE_BY_NAME[prize_name]
    return draw.official.prize_amount(prize_type) or 0


def evaluate_ticket(ticket: Ticket, actual: Draw) -> tuple[str | None, int, int, bool]:
    red_hits = (ticket.red_mask & actual.red_mask).bit_count()
    blue_hit = ticket.blue == actual.blue
    if red_hits == 6 and blue_hit:
        return "一等奖", prize_payout(actual, "一等奖"), red_hits, blue_hit
    if red_hits == 6:
        return "二等奖", prize_payout(actual, "二等奖"), red_hits, blue_hit
    if red_hits == 5 and blue_hit:
        return "三等奖", prize_payout(actual, "三等奖"), red_hits, blue_hit
    if red_hits == 5 or (red_hits == 4 and blue_hit):
        return "四等奖", prize_payout(actual, "四等奖"), red_hits, blue_hit
    if red_hits == 4 or (red_hits == 3 and blue_hit):
        return "五等奖", prize_payout(actual, "五等奖"), red_hits, blue_hit
    if blue_hit:
        return "六等奖", prize_payout(actual, "六等奖"), red_hits, blue_hit
    return None, 0, red_hits, blue_hit


def evaluate_scheme(scheme: Scheme, actual: Draw) -> SchemeResult:
    total_return = 0
    prize_counts: Counter = Counter()
    winning_ticket_count = 0
    value_hit_tickets = 0
    strong_hit_tickets = 0
    max_red_hits = 0
    blue_hit_tickets = 0
    match_scores: list[float] = []
    ticket_returns: list[int] = []
    for ticket in scheme.tickets:
        prize, payout, red_hits, blue_hit = evaluate_ticket(ticket, actual)
        total_return += payout
        ticket_returns.append(payout)
        max_red_hits = max(max_red_hits, red_hits)
        blue_hit_tickets += int(blue_hit)
        match_scores.append(red_hits + (0.5 if blue_hit else 0.0))
        if prize:
            winning_ticket_count += 1
            prize_counts[prize] += 1
            if prize != "六等奖":
                value_hit_tickets += 1
            if prize in {"一等奖", "二等奖", "三等奖", "四等奖"}:
                strong_hit_tickets += 1
    average_match_score = sum(match_scores) / len(match_scores) if match_scores else 0.0
    profit = total_return - scheme.cost
    budget = int(scheme.metadata.get("budget", scheme.cost)) if scheme.metadata else scheme.cost
    reward_rate = total_return / budget if budget else 0.0
    profit_rate = profit / budget if budget else 0.0
    utilization = scheme.cost / budget if budget else 1.0
    utility = (
        0.34 * reward_rate
        + 0.18 * profit_rate
        + 0.12 * utilization
        + 0.20 * (max_red_hits / 6.0)
        + 0.16 * (average_match_score / 6.5)
    )
    return SchemeResult(
        total_return=total_return,
        profit=profit,
        prize_counts=prize_counts,
        winning_ticket_count=winning_ticket_count,
        value_hit_tickets=value_hit_tickets,
        strong_hit_tickets=strong_hit_tickets,
        max_red_hits=max_red_hits,
        blue_hit_tickets=blue_hit_tickets,
        average_match_score=average_match_score,
        utility=utility,
        ticket_returns=ticket_returns,
    )


def build_validation_record(scheme: Scheme, result: SchemeResult) -> ValidationRecord:
    return ValidationRecord(
        cost=scheme.cost,
        total_return=result.total_return,
        winning_ticket_count=result.winning_ticket_count,
        value_hit_tickets=result.value_hit_tickets,
        strong_hit_tickets=result.strong_hit_tickets,
        blue_hit_tickets=result.blue_hit_tickets,
        max_red_hits=result.max_red_hits,
        average_match_score=result.average_match_score,
        ticket_count=len(scheme.tickets),
    )


def serialize_ticket(ticket: Ticket) -> dict[str, object]:
    return {
        "reds": list(ticket.reds),
        "blue": ticket.blue,
        "score": round(ticket.score, 6),
    }


def render_scheme_markdown(scheme: Scheme, budget: int) -> list[str]:
    lines = [
        f"### {budget}元预算",
        f"- 推荐结构：{scheme.description}",
        f"- 实际成本：{scheme.cost} 元",
    ]
    if scheme.family == "hold":
        lines.append("- 动作：本期空仓，等待更强信号。")
    elif scheme.family == "single_pack":
        lines.append(f"- 单式票数：{len(scheme.tickets)} 注")
        preview = scheme.tickets[: min(12, len(scheme.tickets))]
        lines.append("- 前 12 注预览：")
        for ticket in preview:
            red_text = " ".join(f"{number:02d}" for number in ticket.reds)
            lines.append(f"  - {red_text} + {ticket.blue:02d}")
    elif scheme.family == "red_fushi":
        reds = " ".join(f"{number:02d}" for number in scheme.metadata["reds"])
        lines.append(f"- 红球池：{reds}")
        lines.append(f"- 蓝球：{scheme.metadata['blue']:02d}")
    elif scheme.family == "blue_fushi":
        reds = " ".join(f"{number:02d}" for number in scheme.metadata["reds"])
        blues = " ".join(f"{number:02d}" for number in scheme.metadata["blues"])
        lines.append(f"- 红球：{reds}")
        lines.append(f"- 蓝球池：{blues}")
    elif scheme.family == "full_fushi":
        reds = " ".join(f"{number:02d}" for number in scheme.metadata["reds"])
        blues = " ".join(f"{number:02d}" for number in scheme.metadata["blues"])
        lines.append(f"- 红球池：{reds}")
        lines.append(f"- 蓝球池：{blues}")
    elif scheme.family == "dantuo":
        dan = " ".join(f"{number:02d}" for number in scheme.metadata["dan"])
        drag = " ".join(f"{number:02d}" for number in scheme.metadata["drag"])
        blues = " ".join(f"{number:02d}" for number in scheme.metadata["blues"])
        lines.append(f"- 胆码：{dan}")
        lines.append(f"- 拖码：{drag}")
        lines.append(f"- 蓝球池：{blues}")
    return lines


def run_engine(draws: Sequence[Draw]) -> dict[str, object]:
    start_issue = 2020018
    start_index = next(index for index, draw in enumerate(draws) if draw.issue == start_issue)
    red_arm_weights = {arm: 1 / len(ARM_CONFIGS) for arm in ARM_CONFIGS}
    blue_arm_weights = {arm: 1 / len(ARM_CONFIGS) for arm in ARM_CONFIGS}
    red_source_weights = {"heuristic": 0.60, "online_model": 0.40}
    blue_source_weights = {"heuristic": 0.55, "online_model": 0.45}
    red_online_model = OnlineBinaryModel(
        feature_names=FEATURE_NAMES,
        learning_rate=0.28,
        l2=0.002,
        positive_weight=3.5,
    )
    blue_online_model = OnlineBinaryModel(
        feature_names=FEATURE_NAMES,
        learning_rate=0.22,
        l2=0.003,
        positive_weight=6.0,
    )
    ticket_calibrator = TicketReturnCalibrator()
    scheme_gate_calibrator = SchemeGateCalibrator()
    history_snapshots = [
        HistorySnapshot(
            issue=draws[index].issue,
            context_vector=build_history_context_vector(draws[:index]),
            reds=draws[index].reds,
            blue=draws[index].blue,
        )
        for index in range(max(30, 1), start_index)
    ]
    family_stats = {
        budget: {family: {"ema": 0.0, "count": 0} for family in ["single_pack", "red_fushi", "blue_fushi", "full_fushi", "dantuo"]}
        for budget in BUDGETS
    }
    family_live_states = {
        budget: {family: FamilyLiveState() for family in ["single_pack", "red_fushi", "blue_fushi", "full_fushi", "dantuo"]}
        for budget in BUDGETS
    }
    policy_stats = {budget: defaultdict(lambda: {"ema": 0.0, "count": 0}) for budget in BUDGETS}
    profile_stats = {budget: defaultdict(lambda: {"ema": 0.0, "count": 0}) for budget in BUDGETS}
    profile_roi_stats = {budget: defaultdict(lambda: {"ema": 0.0, "count": 0}) for budget in BUDGETS}
    profile_validation_records = {budget: defaultdict(list) for budget in BUDGETS}
    profile_validation_cache = {budget: {} for budget in BUDGETS}
    reward_hit_snapshots = {budget: defaultdict(list) for budget in BUDGETS}
    budget_summaries = {budget: BudgetSummary(budget=budget) for budget in BUDGETS}
    rolling_rows: list[dict[str, object]] = []
    red_arm_quality_totals = {arm: 0.0 for arm in ARM_CONFIGS}
    blue_arm_quality_totals = {arm: 0.0 for arm in ARM_CONFIGS}
    red_source_quality_totals = {source: 0.0 for source in red_source_weights}
    blue_source_quality_totals = {source: 0.0 for source in blue_source_weights}

    for index in range(start_index, len(draws)):
        history = draws[:index]
        actual = draws[index]
        current_context = build_history_context_vector(history)
        red_similarity_scores = build_similarity_feature_map(
            current_context,
            history_snapshots,
            RED_MAX,
            lambda snapshot: snapshot.reds,
            top_k=40,
        )
        blue_similarity_scores = build_similarity_feature_map(
            current_context,
            history_snapshots,
            BLUE_MAX,
            lambda snapshot: (snapshot.blue,),
            top_k=32,
        )
        # ── 混沌 & 概率论特征预计算 ────────────────────────────────────────
        red_hit_getter  = lambda draw, number: number in draw.reds
        blue_hit_getter = lambda draw, number: draw.blue == number
        # 一阶Markov转移矩阵（红球）
        red_markov_trans  = build_markov_transition(history, RED_MAX)
        blue_markov_trans = build_markov_transition(history, BLUE_MAX)
        red_markov_scores  = build_markov_scores(history, RED_MAX,  red_markov_trans)
        blue_markov_scores = build_markov_scores(history, BLUE_MAX, blue_markov_trans)
        # Shannon熵偏差分数
        red_entropy_scores  = build_shannon_entropy_scores(history, RED_MAX,  red_hit_getter)
        blue_entropy_scores = build_shannon_entropy_scores(history, BLUE_MAX, blue_hit_getter)
        # 递归定量分析
        red_recurrence_scores  = build_recurrence_scores(history, RED_MAX,  red_hit_getter)
        blue_recurrence_scores = build_recurrence_scores(history, BLUE_MAX, blue_hit_getter)
        red_features = compute_number_features(
            history,
            RED_MAX,
            red_hit_getter,
            decay_base=0.965,
            similarity_scores=red_similarity_scores,
            markov_scores=red_markov_scores,
            entropy_scores=red_entropy_scores,
            recurrence_scores=red_recurrence_scores,
        )
        blue_features = compute_number_features(
            history,
            BLUE_MAX,
            blue_hit_getter,
            decay_base=0.945,
            similarity_scores=blue_similarity_scores,
            markov_scores=blue_markov_scores,
            entropy_scores=blue_entropy_scores,
            recurrence_scores=blue_recurrence_scores,
        )
        arm_red_scores = {arm: score_arm_numbers(red_features, weights) for arm, weights in ARM_CONFIGS.items()}
        arm_blue_scores = {arm: score_arm_numbers(blue_features, weights) for arm, weights in ARM_CONFIGS.items()}
        heuristic_red_scores = blend_score_maps(red_arm_weights, arm_red_scores)
        heuristic_blue_scores = blend_score_maps(blue_arm_weights, arm_blue_scores)
        online_red_scores = red_online_model.predict_scores(red_features)
        online_blue_scores = blue_online_model.predict_scores(blue_features)
        ensemble_red_scores = blend_score_maps(
            red_source_weights,
            {"heuristic": heuristic_red_scores, "online_model": online_red_scores},
        )
        ensemble_blue_scores = blend_score_maps(
            blue_source_weights,
            {"heuristic": heuristic_blue_scores, "online_model": online_blue_scores},
        )
        shape_context = build_shape_context(history)
        pair_scores, max_pair_score = build_pair_scores(history)
        triplet_scores, max_triplet_score = build_triplet_scores(history)
        shape_context["triplet_scores"] = triplet_scores
        shape_context["max_triplet_score"] = max_triplet_score
        # PageRank 共现网络中心性注入
        _pr = build_pagerank_scores(history)
        shape_context["pagerank_scores"] = _pr
        shape_context["pagerank_max"] = max(_pr.values(), default=1.0)
        # 相位空间吸引子信息注入 shape_context
        attractor_info = build_phase_attractor_score(history)
        shape_context["attractor_correction"] = attractor_info["attractor_correction"]
        shape_context["attractor_density"]    = attractor_info["attractor_density"]
        ranked_reds = sorted(ensemble_red_scores, key=lambda number: (-ensemble_red_scores[number], number))
        ranked_blues = sorted(ensemble_blue_scores, key=lambda number: (-ensemble_blue_scores[number], number))
        red_pool_cache = build_red_pool_cache(ranked_reds, ensemble_red_scores, range(6, 13))
        single_pool = build_candidate_single_tickets(
            ensemble_red_scores,
            ensemble_blue_scores,
            shape_context,
            pair_scores,
            max_pair_score,
            max_pool_size=min(max(BUDGETS) // TICKET_COST, 320),
        )
        regime_labels = build_regime_labels(
            red_arm_weights,
            blue_arm_weights,
            red_source_weights,
            blue_source_weights,
            ensemble_red_scores,
            ensemble_blue_scores,
        )

        schemes_by_budget = {
            budget: build_budget_schemes(
                budget,
                ranked_reds,
                ranked_blues,
                red_pool_cache,
                ensemble_red_scores,
                ensemble_blue_scores,
                shape_context,
                pair_scores,
                max_pair_score,
                single_pool,
                regime_labels,
                ticket_calibrator,
            )
            for budget in BUDGETS
        }

        red_arm_qualities = {arm: red_score_quality(arm_red_scores[arm], actual.reds) for arm in ARM_CONFIGS}
        blue_arm_qualities = {arm: blue_score_quality(arm_blue_scores[arm], actual.blue) for arm in ARM_CONFIGS}
        red_source_qualities = {
            "heuristic": red_score_quality(heuristic_red_scores, actual.reds),
            "online_model": red_score_quality(online_red_scores, actual.reds),
        }
        blue_source_qualities = {
            "heuristic": blue_score_quality(heuristic_blue_scores, actual.blue),
            "online_model": blue_score_quality(online_blue_scores, actual.blue),
        }
        for arm, quality in red_arm_qualities.items():
            red_arm_quality_totals[arm] += quality
        for arm, quality in blue_arm_qualities.items():
            blue_arm_quality_totals[arm] += quality
        for source, quality in red_source_qualities.items():
            red_source_quality_totals[source] += quality
        for source, quality in blue_source_qualities.items():
            blue_source_quality_totals[source] += quality

        calibration_updates: dict[str, dict[tuple[tuple[int, ...], int], tuple[Ticket, int]]] = defaultdict(dict)
        for budget in BUDGETS:
            for state in family_live_states[budget].values():
                if state.cooldown > 0:
                    state.cooldown -= 1
            available_family_candidates = schemes_by_budget[budget]
            available_families = list(available_family_candidates)
            profile_validation_summaries: dict[str, ProfileValidationSummary] = {}
            for family_candidates in available_family_candidates.values():
                for scheme in family_candidates:
                    profile_key = str(scheme.metadata["profile_key"])
                    summary = get_cached_profile_validation_summary(
                        profile_validation_cache[budget],
                        profile_key,
                        budget,
                        profile_validation_records[budget],
                    )
                    profile_validation_summaries[profile_key] = summary
                    annotate_scheme_nested_validation(scheme, summary)
            representative_schemes = {
                family: select_family_scheme(
                    budget,
                    candidates,
                    profile_stats[budget],
                    profile_roi_stats[budget],
                    profile_validation_summaries,
                )
                for family, candidates in available_family_candidates.items()
            }
            validated_family_candidates = {
                family: [
                    scheme
                    for scheme in family_candidates
                    if profile_validation_summaries[str(scheme.metadata["profile_key"])].passed
                ]
                for family, family_candidates in available_family_candidates.items()
            }
            validated_family_candidates = {
                family: family_candidates
                for family, family_candidates in validated_family_candidates.items()
                if family_candidates
            }
            family_stat_scores = normalized_value_map(
                {family: family_stats[budget][family]["ema"] for family in available_families}
            )
            family_proxy_scores = normalized_value_map(
                {family: representative_schemes[family].proxy_score for family in available_families}
            )
            family_gate_rois: dict[str, float] = {}
            red_similarity_peak = max(red_similarity_scores.values(), default=0.0)
            blue_similarity_peak = max(blue_similarity_scores.values(), default=0.0)
            for family, scheme in representative_schemes.items():
                gate_confidence = scheme_gate_confidence(
                    family_stat_scores[family],
                    family_proxy_scores[family],
                    scheme,
                    red_similarity_peak,
                    blue_similarity_peak,
                )
                predicted_gate_roi = scheme_gate_calibrator.predict_roi(budget, family, gate_confidence)
                scheme.metadata["gate_confidence"] = round(gate_confidence, 6)
                scheme.metadata["predicted_gate_roi"] = round(predicted_gate_roi, 6)
                family_gate_rois[family] = predicted_gate_roi
            available_family_candidates = {
                family: limit_family_schemes(
                    family,
                    expand_gate_variants(
                        family,
                        family_candidates,
                        family_gate_rois.get(family, -1.0),
                        ticket_calibrator,
                    ),
                )
                for family, family_candidates in available_family_candidates.items()
            }
            available_family_candidates = {
                family: limit_family_schemes(
                    family,
                    expand_form_variants(
                        family_candidates,
                        profile_stats[budget],
                        profile_roi_stats[budget],
                        ticket_calibrator,
                    ),
                )
                for family, family_candidates in available_family_candidates.items()
            }
            family_hit_alignments = {
                family: compute_reward_hit_alignment(
                    current_context,
                    reward_hit_snapshots[budget][family],
                )
                for family in available_family_candidates
            }
            available_family_candidates = {
                family: limit_family_schemes(
                    family,
                    expand_hit_variants(
                        family_candidates,
                        family_hit_alignments.get(family, 0.0),
                        ticket_calibrator,
                    ),
                )
                for family, family_candidates in available_family_candidates.items()
            }
            profile_validation_summaries = {}
            for family_candidates in available_family_candidates.values():
                for scheme in family_candidates:
                    profile_key = str(scheme.metadata["profile_key"])
                    summary = get_cached_profile_validation_summary(
                        profile_validation_cache[budget],
                        profile_key,
                        budget,
                        profile_validation_records[budget],
                    )
                    profile_validation_summaries[profile_key] = summary
                    annotate_scheme_nested_validation(scheme, summary)
            validated_family_candidates = {
                family: [
                    scheme
                    for scheme in family_candidates
                    if profile_validation_summaries[str(scheme.metadata["profile_key"])].passed
                ]
                for family, family_candidates in available_family_candidates.items()
            }
            validated_family_candidates = {
                family: family_candidates
                for family, family_candidates in validated_family_candidates.items()
                if family_candidates
            }
            blocked_families = {
                family
                for family, state in family_live_states[budget].items()
                if family_is_cooldown_blocked(
                    state,
                    budget,
                    validated_family_candidates.get(family, ()),
                    profile_validation_summaries,
                )
            }
            validated_family_candidates = {
                family: family_candidates
                for family, family_candidates in validated_family_candidates.items()
                if family not in blocked_families
            }
            applicable_policies: list[tuple[float, str, str]] = []
            family_policy_ids: dict[str, list[str]] = {family: [] for family in available_families}
            for family in validated_family_candidates:
                policy_ids = applicable_policy_ids(
                    family,
                    family_stat_scores[family],
                    family_proxy_scores[family],
                )
                family_policy_ids[family] = policy_ids
                for policy_id in policy_ids:
                    state = policy_stats[budget][policy_id]
                    policy_score = policy_selection_score(state["ema"], state["count"])
                    applicable_policies.append((policy_score, family, policy_id))
            chosen_profile_key = "hold"
            if not applicable_policies:
                chosen_family = "hold"
                chosen_policy_id = "nested_validation::hold"
                chosen_scheme = make_hold_scheme(budget, max(family_gate_rois.values(), default=0.0))
                chosen_scheme.metadata["decision"] = "nested_validation_hold"
                chosen_scheme.metadata["validated_family_count"] = 0
                chosen_scheme.metadata["blocked_families"] = sorted(blocked_families)
                best_policy_score = 0.0
            else:
                best_policy_score, chosen_family, chosen_policy_id = max(applicable_policies)
                if best_policy_score <= 0.0:
                    chosen_scheme = make_hold_scheme(budget, family_gate_rois[chosen_family])
                    chosen_family = "hold"
                    chosen_policy_id = "hold"
                    chosen_profile_key = "hold"
                else:
                    family_candidates = validated_family_candidates[chosen_family]
                    chosen_scheme = select_family_scheme(
                        budget,
                        family_candidates,
                        profile_stats[budget],
                        profile_roi_stats[budget],
                        profile_validation_summaries,
                    )
                    chosen_profile_key = str(chosen_scheme.metadata["profile_key"])
            if chosen_family == "hold":
                chosen_profile_key = "hold"
                chosen_scheme.metadata["nested_validation_required"] = True
                if chosen_scheme.metadata.get("decision") == "nested_validation_hold":
                    chosen_scheme.metadata["validated_profiles"] = sorted(
                        profile_key
                        for profile_key, summary in profile_validation_summaries.items()
                        if summary.passed
                    )
                    chosen_scheme.metadata["blocked_families"] = sorted(blocked_families)
            else:
                chosen_scheme.metadata["nested_validation_required"] = True
            if chosen_family != "hold" and not profile_validation_summaries[chosen_profile_key].passed:
                chosen_scheme = make_hold_scheme(budget, family_gate_rois.get(chosen_family, 0.0))
                chosen_family = "hold"
                chosen_policy_id = f"{chosen_policy_id}|nested_validation_fail"
                chosen_profile_key = "hold"
            chosen_result = evaluate_scheme(chosen_scheme, actual)
            if chosen_family != "hold":
                update_family_live_state(
                    family_live_states[budget][chosen_family],
                    budget,
                    chosen_scheme,
                    chosen_result,
                )
            if chosen_family != "hold" and chosen_scheme.cost > 0 and chosen_result.total_return > 0:
                reward_hit_snapshots[budget][chosen_family].append(
                    RewardHitSnapshot(
                        issue=actual.issue,
                        budget=budget,
                        family=chosen_family,
                        profile_key=chosen_profile_key,
                        context_vector=current_context.copy(),
                        reward_ratio=min(chosen_result.total_return / chosen_scheme.cost, 3.0),
                        profit_ratio=chosen_result.profit / chosen_scheme.cost,
                    )
                )
            budget_summaries[budget].add(chosen_family, chosen_scheme, chosen_result)
            rolling_rows.append(
                {
                    "issue": actual.issue,
                    "budget": budget,
                    "chosen_family": chosen_family,
                    "cost": chosen_scheme.cost,
                    "family_stat_score": round(family_stat_scores.get(chosen_family, 0.0), 4),
                    "family_proxy_score": round(family_proxy_scores.get(chosen_family, 0.0), 4),
                    "expected_profit_ratio": chosen_scheme.metadata.get("expected_profit_ratio", ""),
                    "predicted_gate_roi": chosen_scheme.metadata.get("predicted_gate_roi", ""),
                    "gate_confidence": chosen_scheme.metadata.get("gate_confidence", ""),
                    "chosen_policy": chosen_policy_id,
                    "policy_score": round(best_policy_score, 6),
                    "chosen_profile": chosen_profile_key,
                    "nested_validation_pass": chosen_scheme.metadata.get("nested_validation_pass", False),
                    "nested_validation_mode": chosen_scheme.metadata.get("nested_validation_mode", ""),
                    "nested_validation_score": chosen_scheme.metadata.get("nested_validation_score", ""),
                    "nested_validation_reason": chosen_scheme.metadata.get("nested_validation_reason", ""),
                    "hit_validation_pass": chosen_scheme.metadata.get("hit_validation_pass", False),
                    "hit_validation_score": chosen_scheme.metadata.get("hit_validation_score", ""),
                    "winning_issue": int(chosen_result.winning_ticket_count > 0),
                    "value_hit_issue": int(chosen_result.value_hit_tickets > 0),
                    "strong_hit_issue": int(chosen_result.strong_hit_tickets > 0),
                    "winning_ticket_count": chosen_result.winning_ticket_count,
                    "total_return": chosen_result.total_return,
                    "total_profit": chosen_result.profit,
                    "max_red_hits": chosen_result.max_red_hits,
                    "blue_hit_tickets": chosen_result.blue_hit_tickets,
                    "avg_match_score": round(chosen_result.average_match_score, 4),
                    "top_red_source": max(red_source_weights, key=red_source_weights.get),
                    "top_blue_source": max(blue_source_weights, key=blue_source_weights.get),
                    "red_similarity_peak": round(red_similarity_peak, 4),
                    "blue_similarity_peak": round(blue_similarity_peak, 4),
                    "actual_reds": " ".join(f"{number:02d}" for number in actual.reds),
                    "actual_blue": f"{actual.blue:02d}",
                }
            )
            profile_result_cache: dict[str, SchemeResult] = {}
            for family, family_candidates in available_family_candidates.items():
                representative_scheme = representative_schemes[family]
                representative_profile = str(representative_scheme.metadata["profile_key"])
                if family == chosen_family and chosen_profile_key == representative_profile:
                    representative_result = chosen_result
                else:
                    representative_result = evaluate_scheme(representative_scheme, actual)
                state = family_stats[budget][family]
                reward_signal = scheme_reward_signal(representative_result, representative_scheme)
                state["ema"] = reward_signal if state["count"] == 0 else 0.92 * state["ema"] + 0.08 * reward_signal
                state["count"] += 1
                scheme_gate_calibrator.update(
                    budget=budget,
                    family=family,
                    confidence=float(representative_scheme.metadata.get("gate_confidence", 0.0)),
                    roi=(representative_result.profit / representative_scheme.cost) if representative_scheme.cost else 0.0,
                )
                for policy_id in family_policy_ids[family]:
                    policy_state = policy_stats[budget][policy_id]
                    policy_roi = reward_signal
                    policy_state["ema"] = (
                        policy_roi if policy_state["count"] == 0 else 0.90 * policy_state["ema"] + 0.10 * policy_roi
                    )
                    policy_state["count"] += 1
                for scheme in family_candidates:
                    profile_key = str(scheme.metadata["profile_key"])
                    if family == chosen_family and profile_key == chosen_profile_key:
                        result = chosen_result
                    elif profile_key in profile_result_cache:
                        result = profile_result_cache[profile_key]
                    else:
                        result = evaluate_scheme(scheme, actual)
                        profile_result_cache[profile_key] = result
                    profile_state = profile_stats[budget][profile_key]
                    profile_signal = scheme_reward_signal(result, scheme)
                    profile_state["ema"] = (
                        profile_signal if profile_state["count"] == 0 else 0.90 * profile_state["ema"] + 0.10 * profile_signal
                    )
                    profile_state["count"] += 1
                    roi_state = profile_roi_stats[budget][profile_key]
                    roi_signal = realized_roi_signal(result, scheme)
                    roi_state["ema"] = roi_signal if roi_state["count"] == 0 else 0.90 * roi_state["ema"] + 0.10 * roi_signal
                    roi_state["count"] += 1
                    profile_validation_records[budget][profile_key].append(build_validation_record(scheme, result))
                    for ticket, payout in zip(scheme.tickets, result.ticket_returns):
                        ticket_key = (ticket.reds, ticket.blue)
                        if ticket_key not in calibration_updates[family]:
                            calibration_updates[family][ticket_key] = (ticket, payout)

        for family, ticket_map in calibration_updates.items():
            for ticket, payout in ticket_map.values():
                ticket_calibrator.update_ticket(family, ticket, payout)

        red_arm_weights = update_weights(red_arm_weights, red_arm_qualities, eta=0.50)
        blue_arm_weights = update_weights(blue_arm_weights, blue_arm_qualities, eta=0.50)
        red_source_weights = update_weights(red_source_weights, red_source_qualities, eta=0.35)
        blue_source_weights = update_weights(blue_source_weights, blue_source_qualities, eta=0.35)
        red_online_model.update(red_features, set(actual.reds))
        blue_online_model.update(blue_features, {actual.blue})
        history_snapshots.append(
            HistorySnapshot(
                issue=actual.issue,
                context_vector=current_context.copy(),
                reds=actual.reds,
                blue=actual.blue,
            )
        )

    full_history = draws
    final_context = build_history_context_vector(full_history)
    final_red_similarity_scores = build_similarity_feature_map(
        final_context,
        history_snapshots,
        RED_MAX,
        lambda snapshot: snapshot.reds,
        top_k=40,
    )
    final_blue_similarity_scores = build_similarity_feature_map(
        final_context,
        history_snapshots,
        BLUE_MAX,
        lambda snapshot: (snapshot.blue,),
        top_k=32,
    )
    # ── 混沌 & 概率论特征预计算（最终预测段） ──────────────────────────────
    final_red_hit_getter  = lambda draw, number: number in draw.reds
    final_blue_hit_getter = lambda draw, number: draw.blue == number
    final_red_markov_trans  = build_markov_transition(full_history, RED_MAX)
    final_blue_markov_trans = build_markov_transition(full_history, BLUE_MAX)
    final_red_markov_scores  = build_markov_scores(full_history, RED_MAX,  final_red_markov_trans)
    final_blue_markov_scores = build_markov_scores(full_history, BLUE_MAX, final_blue_markov_trans)
    final_red_entropy_scores  = build_shannon_entropy_scores(full_history, RED_MAX,  final_red_hit_getter)
    final_blue_entropy_scores = build_shannon_entropy_scores(full_history, BLUE_MAX, final_blue_hit_getter)
    final_red_recurrence_scores  = build_recurrence_scores(full_history, RED_MAX,  final_red_hit_getter)
    final_blue_recurrence_scores = build_recurrence_scores(full_history, BLUE_MAX, final_blue_hit_getter)
    final_red_features = compute_number_features(
        full_history,
        RED_MAX,
        final_red_hit_getter,
        decay_base=0.965,
        similarity_scores=final_red_similarity_scores,
        markov_scores=final_red_markov_scores,
        entropy_scores=final_red_entropy_scores,
        recurrence_scores=final_red_recurrence_scores,
    )
    final_blue_features = compute_number_features(
        full_history,
        BLUE_MAX,
        final_blue_hit_getter,
        decay_base=0.945,
        similarity_scores=final_blue_similarity_scores,
        markov_scores=final_blue_markov_scores,
        entropy_scores=final_blue_entropy_scores,
        recurrence_scores=final_blue_recurrence_scores,
    )
    final_arm_red_scores = {arm: score_arm_numbers(final_red_features, weights) for arm, weights in ARM_CONFIGS.items()}
    final_arm_blue_scores = {arm: score_arm_numbers(final_blue_features, weights) for arm, weights in ARM_CONFIGS.items()}
    final_heuristic_red_scores = blend_score_maps(red_arm_weights, final_arm_red_scores)
    final_heuristic_blue_scores = blend_score_maps(blue_arm_weights, final_arm_blue_scores)
    final_online_red_scores = red_online_model.predict_scores(final_red_features)
    final_online_blue_scores = blue_online_model.predict_scores(final_blue_features)
    final_ensemble_red_scores = blend_score_maps(
        red_source_weights,
        {"heuristic": final_heuristic_red_scores, "online_model": final_online_red_scores},
    )
    final_ensemble_blue_scores = blend_score_maps(
        blue_source_weights,
        {"heuristic": final_heuristic_blue_scores, "online_model": final_online_blue_scores},
    )
    final_shape_context = build_shape_context(full_history)
    final_pair_scores, final_max_pair_score = build_pair_scores(full_history)
    final_triplet_scores, final_max_triplet_score = build_triplet_scores(full_history)
    final_shape_context["triplet_scores"] = final_triplet_scores
    final_shape_context["max_triplet_score"] = final_max_triplet_score
    # PageRank 共现网络中心性注入（最终预测段）
    _final_pr = build_pagerank_scores(full_history)
    final_shape_context["pagerank_scores"] = _final_pr
    final_shape_context["pagerank_max"] = max(_final_pr.values(), default=1.0)
    # 相位空间吸引子（最终预测段）
    final_attractor_info = build_phase_attractor_score(full_history)
    final_shape_context["attractor_correction"] = final_attractor_info["attractor_correction"]
    final_shape_context["attractor_density"]    = final_attractor_info["attractor_density"]
    final_ranked_reds = sorted(final_ensemble_red_scores, key=lambda number: (-final_ensemble_red_scores[number], number))
    final_ranked_blues = sorted(final_ensemble_blue_scores, key=lambda number: (-final_ensemble_blue_scores[number], number))
    final_red_pool_cache = build_red_pool_cache(final_ranked_reds, final_ensemble_red_scores, range(6, 13))
    final_single_pool = build_candidate_single_tickets(
        final_ensemble_red_scores,
        final_ensemble_blue_scores,
        final_shape_context,
        final_pair_scores,
        final_max_pair_score,
        max_pool_size=min(max(BUDGETS) // TICKET_COST, 320),
    )
    final_regime_labels = build_regime_labels(
        red_arm_weights,
        blue_arm_weights,
        red_source_weights,
        blue_source_weights,
        final_ensemble_red_scores,
        final_ensemble_blue_scores,
    )
    final_schemes = {
        budget: build_budget_schemes(
            budget,
            final_ranked_reds,
            final_ranked_blues,
            final_red_pool_cache,
            final_ensemble_red_scores,
            final_ensemble_blue_scores,
            final_shape_context,
            final_pair_scores,
            final_max_pair_score,
            final_single_pool,
            final_regime_labels,
            ticket_calibrator,
        )
        for budget in BUDGETS
    }
    final_predictions = {}
    final_validation_diagnostics: dict[int, list[dict[str, object]]] = {}
    for budget in BUDGETS:
        final_family_candidates = final_schemes[budget]
        final_profile_validation_summaries: dict[str, ProfileValidationSummary] = {}
        for family_candidates in final_family_candidates.values():
            for scheme in family_candidates:
                profile_key = str(scheme.metadata["profile_key"])
                summary = get_cached_profile_validation_summary(
                    profile_validation_cache[budget],
                    profile_key,
                    budget,
                    profile_validation_records[budget],
                )
                final_profile_validation_summaries[profile_key] = summary
                annotate_scheme_nested_validation(scheme, summary)
        final_validation_diagnostics[budget] = sorted(
            [
                {
                    "profile": profile_key,
                    "passed": summary.passed,
                    "score": round(summary.score, 6),
                    "mode": summary.mode,
                    "observations": summary.roi.observations,
                    "overall_roi": round(summary.roi.overall_roi, 6),
                    "recent_roi": round(summary.roi.recent_roi, 6),
                    "pass_ratio": round(summary.roi.pass_ratio, 6),
                    "overall_profitable_hits": summary.roi.overall_profitable_hits,
                    "recent_profitable_hits": summary.roi.recent_profitable_hits,
                    "positive_return_concentration": round(summary.roi.positive_return_concentration, 6),
                    "overall_hit_score": round(summary.hit.overall_hit_score, 6),
                    "recent_hit_score": round(summary.hit.recent_hit_score, 6),
                    "overall_any_hit_rate": round(summary.hit.overall_any_hit_rate, 6),
                    "recent_any_hit_rate": round(summary.hit.recent_any_hit_rate, 6),
                    "reason": summary.reason,
                    "roi_reason": summary.roi.reason,
                    "hit_reason": summary.hit.reason,
                }
                for profile_key, summary in final_profile_validation_summaries.items()
            ],
            key=lambda item: (
                int(bool(item["passed"])),
                float(item["score"]),
                float(item["overall_hit_score"]),
                float(item["overall_roi"]),
                -float(item["positive_return_concentration"]),
            ),
            reverse=True,
        )[:30]
        final_representative_schemes = {
            family: select_family_scheme(
                budget,
                candidates,
                profile_stats[budget],
                profile_roi_stats[budget],
                final_profile_validation_summaries,
            )
            for family, candidates in final_family_candidates.items()
        }
        final_validated_family_candidates = {
            family: [
                scheme
                for scheme in family_candidates
                if final_profile_validation_summaries[str(scheme.metadata["profile_key"])].passed
            ]
            for family, family_candidates in final_family_candidates.items()
        }
        final_validated_family_candidates = {
            family: family_candidates
            for family, family_candidates in final_validated_family_candidates.items()
            if family_candidates
        }
        final_family_stat_scores = normalized_value_map(
            {family: family_stats[budget][family]["ema"] for family in final_family_candidates}
        )
        final_family_proxy_scores = normalized_value_map(
            {family: final_representative_schemes[family].proxy_score for family in final_family_candidates}
        )
        final_gate_rois: dict[str, float] = {}
        final_red_similarity_peak = max(final_red_similarity_scores.values(), default=0.0)
        final_blue_similarity_peak = max(final_blue_similarity_scores.values(), default=0.0)
        for family_name, scheme in final_representative_schemes.items():
            gate_confidence = scheme_gate_confidence(
                final_family_stat_scores[family_name],
                final_family_proxy_scores[family_name],
                scheme,
                final_red_similarity_peak,
                final_blue_similarity_peak,
            )
            predicted_gate_roi = scheme_gate_calibrator.predict_roi(budget, family_name, gate_confidence)
            scheme.metadata["gate_confidence"] = round(gate_confidence, 6)
            scheme.metadata["predicted_gate_roi"] = round(predicted_gate_roi, 6)
            final_gate_rois[family_name] = predicted_gate_roi
        final_family_candidates = {
            family_name: limit_family_schemes(
                family_name,
                expand_gate_variants(
                    family_name,
                    family_candidates,
                    final_gate_rois.get(family_name, -1.0),
                    ticket_calibrator,
                ),
            )
            for family_name, family_candidates in final_family_candidates.items()
        }
        final_family_candidates = {
            family_name: limit_family_schemes(
                family_name,
                expand_form_variants(
                    family_candidates,
                    profile_stats[budget],
                    profile_roi_stats[budget],
                    ticket_calibrator,
                ),
            )
            for family_name, family_candidates in final_family_candidates.items()
        }
        final_family_hit_alignments = {
            family_name: compute_reward_hit_alignment(
                final_context,
                reward_hit_snapshots[budget][family_name],
            )
            for family_name in final_family_candidates
        }
        final_family_candidates = {
            family_name: limit_family_schemes(
                family_name,
                expand_hit_variants(
                    family_candidates,
                    final_family_hit_alignments.get(family_name, 0.0),
                    ticket_calibrator,
                ),
            )
            for family_name, family_candidates in final_family_candidates.items()
        }
        final_profile_validation_summaries = {}
        for family_candidates in final_family_candidates.values():
            for scheme in family_candidates:
                profile_key = str(scheme.metadata["profile_key"])
                summary = get_cached_profile_validation_summary(
                    profile_validation_cache[budget],
                    profile_key,
                    budget,
                    profile_validation_records[budget],
                )
                final_profile_validation_summaries[profile_key] = summary
                annotate_scheme_nested_validation(scheme, summary)
        final_validation_diagnostics[budget] = sorted(
            [
                {
                    "profile": profile_key,
                    "passed": summary.passed,
                    "score": round(summary.score, 6),
                    "mode": summary.mode,
                    "observations": summary.roi.observations,
                    "overall_roi": round(summary.roi.overall_roi, 6),
                    "recent_roi": round(summary.roi.recent_roi, 6),
                    "pass_ratio": round(summary.roi.pass_ratio, 6),
                    "overall_profitable_hits": summary.roi.overall_profitable_hits,
                    "recent_profitable_hits": summary.roi.recent_profitable_hits,
                    "positive_return_concentration": round(summary.roi.positive_return_concentration, 6),
                    "overall_hit_score": round(summary.hit.overall_hit_score, 6),
                    "recent_hit_score": round(summary.hit.recent_hit_score, 6),
                    "overall_any_hit_rate": round(summary.hit.overall_any_hit_rate, 6),
                    "recent_any_hit_rate": round(summary.hit.recent_any_hit_rate, 6),
                    "reason": summary.reason,
                    "roi_reason": summary.roi.reason,
                    "hit_reason": summary.hit.reason,
                }
                for profile_key, summary in final_profile_validation_summaries.items()
            ],
            key=lambda item: (
                int(bool(item["passed"])),
                float(item["score"]),
                float(item["overall_hit_score"]),
                float(item["overall_roi"]),
                -float(item["positive_return_concentration"]),
            ),
            reverse=True,
        )[:30]
        final_validated_family_candidates = {
            family: [
                scheme
                for scheme in family_candidates
                if final_profile_validation_summaries[str(scheme.metadata["profile_key"])].passed
            ]
            for family, family_candidates in final_family_candidates.items()
        }
        final_validated_family_candidates = {
            family: family_candidates
            for family, family_candidates in final_validated_family_candidates.items()
            if family_candidates
        }
        final_blocked_families = {
            family_name
            for family_name, state in family_live_states[budget].items()
            if family_is_cooldown_blocked(
                state,
                budget,
                final_validated_family_candidates.get(family_name, ()),
                final_profile_validation_summaries,
            )
        }
        final_validated_family_candidates = {
            family: family_candidates
            for family, family_candidates in final_validated_family_candidates.items()
            if family not in final_blocked_families
        }
        final_applicable_policies: list[tuple[float, str, str]] = []
        for family_name in final_validated_family_candidates:
            for policy_id in applicable_policy_ids(
                family_name,
                final_family_stat_scores[family_name],
                final_family_proxy_scores[family_name],
            ):
                state = policy_stats[budget][policy_id]
                policy_score = policy_selection_score(state["ema"], state["count"])
                final_applicable_policies.append((policy_score, family_name, policy_id))
        chosen_profile_key = "hold"
        if not final_applicable_policies:
            family = "hold"
            chosen_policy_id = "nested_validation::hold"
            chosen_profile_key = "hold"
            selected_scheme = make_hold_scheme(budget, max(final_gate_rois.values(), default=0.0))
            selected_scheme.metadata["decision"] = "nested_validation_hold"
            selected_scheme.metadata["nested_validation_required"] = True
            selected_scheme.metadata["validated_family_count"] = 0
            selected_scheme.metadata["blocked_families"] = sorted(final_blocked_families)
            selected_scheme.metadata["validated_profiles"] = sorted(
                profile_key
                for profile_key, summary in final_profile_validation_summaries.items()
                if summary.passed
            )
            best_policy_score = 0.0
        else:
            best_policy_score, family, chosen_policy_id = max(final_applicable_policies)
            if best_policy_score <= 0.0:
                selected_scheme = make_hold_scheme(budget, final_gate_rois[family])
                family = "hold"
                chosen_policy_id = "hold"
                chosen_profile_key = "hold"
            else:
                family_candidates = final_validated_family_candidates[family]
                selected_scheme = select_family_scheme(
                    budget,
                    family_candidates,
                    profile_stats[budget],
                    profile_roi_stats[budget],
                    final_profile_validation_summaries,
                )
                chosen_profile_key = str(selected_scheme.metadata["profile_key"])
        selected_scheme.metadata["nested_validation_required"] = True
        final_predictions[budget] = {
            "selected_family": family,
            "family_stat_ema": round(family_stats[budget].get(family, {"ema": 0.0})["ema"], 6),
            "family_stat_score": round(final_family_stat_scores.get(family, 0.0), 6),
            "family_proxy_score": round(final_family_proxy_scores.get(family, 0.0), 6),
            "predicted_gate_roi": round(final_gate_rois.get(family, selected_scheme.metadata.get("predicted_gate_roi", 0.0)), 6),
            "chosen_policy": chosen_policy_id,
            "policy_score": round(best_policy_score, 6),
            "chosen_profile": chosen_profile_key,
            "scheme": selected_scheme,
        }

    return {
        "budget_summaries": budget_summaries,
        "rolling_rows": rolling_rows,
        "final_red_arm_weights": red_arm_weights,
        "final_blue_arm_weights": blue_arm_weights,
        "final_red_source_weights": red_source_weights,
        "final_blue_source_weights": blue_source_weights,
        "red_model_coefficients": {
            feature_name: round(float(weight), 6)
            for feature_name, weight in zip(FEATURE_NAMES, red_online_model.weights)
        },
        "blue_model_coefficients": {
            feature_name: round(float(weight), 6)
            for feature_name, weight in zip(FEATURE_NAMES, blue_online_model.weights)
        },
        "average_red_arm_quality": {
            arm: round(total / (len(draws) - start_index), 6) for arm, total in red_arm_quality_totals.items()
        },
        "average_blue_arm_quality": {
            arm: round(total / (len(draws) - start_index), 6) for arm, total in blue_arm_quality_totals.items()
        },
        "average_red_source_quality": {
            source: round(total / (len(draws) - start_index), 6) for source, total in red_source_quality_totals.items()
        },
        "average_blue_source_quality": {
            source: round(total / (len(draws) - start_index), 6) for source, total in blue_source_quality_totals.items()
        },
        "ticket_calibration_samples": ticket_calibrator.overall_stat.count,
        "gate_budget_roi_means": {
            budget: round(scheme_gate_calibrator.budget_overall_stats[budget].mean, 6)
            for budget in BUDGETS
            if scheme_gate_calibrator.budget_overall_stats[budget].count
        },
        "similarity_snapshot_count": len(history_snapshots),
        "final_red_scores": final_ensemble_red_scores,
        "final_blue_scores": final_ensemble_blue_scores,
        "family_stats": family_stats,
        "family_live_states": {
            budget: {
                family: {
                    "count": state.count,
                    "ema_profit": round(state.ema_profit, 6),
                    "ema_return": round(state.ema_return, 6),
                    "ema_hit": round(state.ema_hit, 6),
                    "ema_value_hit": round(state.ema_value_hit, 6),
                    "loss_streak": state.loss_streak,
                    "hitless_streak": state.hitless_streak,
                    "cooldown": state.cooldown,
                }
                for family, state in states.items()
            }
            for budget, states in family_live_states.items()
        },
        "final_predictions": final_predictions,
        "final_validation_diagnostics": final_validation_diagnostics,
    }


def write_csv(path: Path, rows: Iterable[dict[str, object]]) -> None:
    rows = list(rows)
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def summarize_top_reward_issues(rolling_rows: Sequence[dict[str, object]], top_n: int = 20) -> list[dict[str, object]]:
    issue_stats: dict[int, dict[str, object]] = {}
    for row in rolling_rows:
        chosen_family = str(row["chosen_family"])
        total_return = int(row["total_return"])
        if chosen_family == "hold" or total_return <= 0:
            continue
        issue = int(row["issue"])
        stat = issue_stats.setdefault(
            issue,
            {
                "issue": issue,
                "total_return": 0,
                "total_cost": 0,
                "total_profit": 0,
                "win_count": 0,
                "hits": [],
            },
        )
        cost = int(row["cost"])
        total_profit = int(row["total_profit"])
        stat["total_return"] += total_return
        stat["total_cost"] += cost
        stat["total_profit"] += total_profit
        stat["win_count"] += 1
        stat["hits"].append(
            {
                "budget": int(row["budget"]),
                "family": chosen_family,
                "profile": str(row["chosen_profile"]),
                "return": total_return,
                "profit": total_profit,
                "reason": str(row["nested_validation_reason"]),
            }
        )
    ranked = sorted(
        issue_stats.values(),
        key=lambda item: (int(item["total_return"]), int(item["win_count"]), int(item["total_profit"])),
        reverse=True,
    )[:top_n]
    for item in ranked:
        total_cost = int(item["total_cost"])
        item["roi"] = round((int(item["total_return"]) / total_cost) - 1.0, 6) if total_cost else 0.0
    return ranked


def main() -> None:
    parser = argparse.ArgumentParser(description="双色球自适应预测与预算组合引擎")
    parser.add_argument("--excel", required=True, type=Path, help="双色球开奖 Excel 路径")
    parser.add_argument("--rule-doc", required=True, type=Path, help="双色球规则 DOCX 路径")
    parser.add_argument("--output-dir", required=True, type=Path, help="输出目录")
    parser.add_argument("--no-official", action="store_true", help="跳过官方接口抓取，仅使用本地缓存或 Excel 数据")
    args = parser.parse_args()

    raw_draws = load_draws(args.excel)
    rule_text = extract_rule_text(args.rule_doc)
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    official_cache_path = output_dir / "official_draws.json"

    official_sync: dict[str, object]
    if args.no_official:
        # 离线模式：只使用本地缓存，如果缓存不存在则只用 Excel
        if official_cache_path.exists():
            import json as _json
            official_records = _json.loads(official_cache_path.read_text(encoding="utf-8"))
            draws, official_sync = merge_draws_with_official(raw_draws, official_records)
            official_sync["official_source"] = "cache_only"
        else:
            draws = raw_draws
            official_sync = {
                "official_source": "excel_only",
                "official_record_count": 0,
                "official_coverage_on_excel": 0,
                "official_coverage_ratio_on_excel": 0.0,
                "appended_latest_issue_count": 0,
                "latest_issue_after_merge": draws[-1].issue if draws else None,
                "mismatch_issues": [],
            }
    else:
        try:
            official_records, official_source = load_or_refresh_ssq_cache(
                cache_path=official_cache_path,
                issue_start=raw_draws[0].issue,
            )
            draws, official_sync = merge_draws_with_official(raw_draws, official_records)
            official_sync["official_source"] = official_source
        except Exception as exc:
            draws = raw_draws
            official_sync = {
                "official_source": "unavailable",
                "official_record_count": 0,
                "official_coverage_on_excel": 0,
                "official_coverage_ratio_on_excel": 0.0,
                "appended_latest_issue_count": 0,
                "latest_issue_after_merge": draws[-1].issue if draws else None,
                "mismatch_issues": [],
                "error": str(exc),
            }

    engine = run_engine(draws)
    next_issue = infer_next_issue(draws)
    next_date = infer_next_draw_date(draws[-1].draw_date)

    budget_summary_records = [summary.to_record() for summary in engine["budget_summaries"].values()]
    rolling_rows = engine["rolling_rows"]
    top_reward_issue_records = summarize_top_reward_issues(rolling_rows)
    final_predictions_payload = {}
    prediction_ticket_rows: list[dict[str, object]] = []
    for budget, prediction in engine["final_predictions"].items():
        scheme: Scheme = prediction["scheme"]
        final_predictions_payload[str(budget)] = {
            "selected_family": prediction["selected_family"],
            "family_stat_ema": prediction["family_stat_ema"],
            "family_stat_score": prediction["family_stat_score"],
            "family_proxy_score": prediction["family_proxy_score"],
            "predicted_gate_roi": prediction.get("predicted_gate_roi"),
            "chosen_policy": prediction.get("chosen_policy"),
            "policy_score": prediction.get("policy_score"),
            "chosen_profile": prediction.get("chosen_profile"),
            "description": scheme.description,
            "cost": scheme.cost,
            "metadata": scheme.metadata,
            "tickets": [serialize_ticket(ticket) for ticket in scheme.tickets],
        }
        for ticket_index, ticket in enumerate(scheme.tickets, start=1):
            prediction_ticket_rows.append(
                {
                    "budget": budget,
                    "family": prediction["selected_family"],
                    "description": scheme.description,
                    "ticket_index": ticket_index,
                    "reds": " ".join(f"{number:02d}" for number in ticket.reds),
                    "blue": f"{ticket.blue:02d}",
                    "score": round(ticket.score, 6),
                }
            )

    write_csv(output_dir / "rolling_backtest.csv", rolling_rows)
    write_csv(output_dir / "budget_summary.csv", budget_summary_records)
    write_csv(output_dir / "prediction_tickets.csv", prediction_ticket_rows)
    write_json(output_dir / "budget_summary.json", budget_summary_records)
    write_json(output_dir / "prediction_next_issue.json", final_predictions_payload)
    write_json(output_dir / "top_reward_issues.json", top_reward_issue_records)
    engine_state_payload = {
        "final_red_arm_weights": engine["final_red_arm_weights"],
        "final_blue_arm_weights": engine["final_blue_arm_weights"],
        "final_red_source_weights": engine["final_red_source_weights"],
        "final_blue_source_weights": engine["final_blue_source_weights"],
        "red_model_coefficients": engine["red_model_coefficients"],
        "blue_model_coefficients": engine["blue_model_coefficients"],
        "average_red_arm_quality": engine["average_red_arm_quality"],
        "average_blue_arm_quality": engine["average_blue_arm_quality"],
        "average_red_source_quality": engine["average_red_source_quality"],
        "average_blue_source_quality": engine["average_blue_source_quality"],
        "ticket_calibration_samples": engine["ticket_calibration_samples"],
        "gate_budget_roi_means": engine.get("gate_budget_roi_means", {}),
        "similarity_snapshot_count": engine["similarity_snapshot_count"],
        "top_reds": sorted(engine["final_red_scores"], key=engine["final_red_scores"].get, reverse=True)[:15],
        "top_blues": sorted(engine["final_blue_scores"], key=engine["final_blue_scores"].get, reverse=True)[:8],
        "family_stats": engine["family_stats"],
        "family_live_states": engine["family_live_states"],
        "final_validation_diagnostics": engine["final_validation_diagnostics"],
        "top_reward_issues": top_reward_issue_records[:10],
        "official_sync": official_sync,
    }
    write_json(output_dir / "engine_state.json", engine_state_payload)
    write_json(output_dir / "adaptive_diagnostics.json", engine_state_payload)
    write_json(output_dir / "official_enrichment.json", official_sync)

    report_lines = [
        "# 双色球自适应回测与预测结果",
        "",
        "## 说明",
        "- 数据源：附件1《双色球最近1000期开奖结果.xlsx》",
        "- 官方增强源：中国福彩网双色球往期开奖接口（https://www.cwl.gov.cn/ygkj/wqkjgg/ssq/）",
        "- 规则源：附件2《福利彩票双色球游戏规则.docx》",
        f"- 历史范围：{draws[0].issue} 至 {draws[-1].issue}，共 {len(draws)} 期",
        "- 初始训练窗口：2019069 至 2020017，共 100 期",
        f"- 首个滚动预测目标：2020018",
        f"- 最终预测目标：{next_issue}（推断开奖日 {next_date}）",
        "- 中奖金额口径：若当期已匹配到中国福彩网官方奖级金额，则一等奖至六等奖全部按当期官方单注奖金精确回测；缺失时才回退到规则固定奖金。",
        "- 风险提示：双色球开奖本质上接近独立随机事件，以下结果只能作为基于历史拟合的预算优化参考，不能视为确定中奖承诺。",
        "",
        "## 引擎进化点",
        "- 红球与蓝球拆分为两套独立在线权重，不再共用同一组策略权重。",
        "- 启发式策略层之外，新增在线二分类模型层，并与启发式层做自适应融合。",
        "- 新增相似期检索特征，用历史上相近局面的后续开奖来增强当前号码打分。",
        "- 相似期上下文已接入官方销售额、奖池、一二等奖注数/金额、一等奖地区分布与固定奖负担等市场维度。",
        "- 预算结构选择改为“历史表现 EMA + 期望回报校准分 + 探索项”共同决定。",
        "",
        "## 规则摘录",
        "- 单注 2 元；红球从 1-33 选 6 个，蓝球从 1-16 选 1 个。",
        "- 三等奖：5 红 + 1 蓝，3000 元。",
        "- 四等奖：5 红 或 4 红 + 1 蓝，200 元。",
        "- 五等奖：4 红 或 3 红 + 1 蓝，10 元。",
        "- 六等奖：1 蓝，5 元。",
        "",
        "## 回测摘要",
    ]
    for record in budget_summary_records:
        report_lines.append(
            f"- {record['budget']}元：全奖级 ROI={record['roi_all_prizes']}, "
            f"总成本={record['total_cost']}, 奖金回收={record['total_return_all_prizes']}, "
            f"中奖期命中率={record['winning_issue_rate']}, "
            f"五等奖及以上命中率={record['value_hit_issue_rate']}, "
            f"四等奖及以上命中率={record['strong_hit_issue_rate']}, "
            f"常用结构={record['family_usage']}"
        )
    report_lines.extend(["", "## 高回报命中期数"])
    if top_reward_issue_records:
        for item in top_reward_issue_records[:10]:
            hit_preview = "; ".join(
                f"{hit['budget']}元 {hit['family']} {hit['profile']} 回收{hit['return']}"
                for hit in item["hits"][:3]
            )
            report_lines.append(
                f"- {item['issue']}期：回收={item['total_return']}，成本={item['total_cost']}，"
                f"收益={item['total_profit']}，ROI={item['roi']}，命中方案={hit_preview}"
            )
    else:
        report_lines.append("- 当前滚动回测中暂无正回收期数。")
    report_lines.extend(
        [
            "",
            "## 官方同步情况",
            f"- 官方记录数：{official_sync['official_record_count']}，Excel 覆盖率：{official_sync['official_coverage_ratio_on_excel']}",
            f"- 追加到训练集的新官方期数：{official_sync['appended_latest_issue_count']}，当前已训练到：{official_sync['latest_issue_after_merge']}",
            f"- 官方数据来源状态：{official_sync['official_source']}",
            "",
            "## 最终推荐",
        ]
    )
    for budget in BUDGETS:
        scheme: Scheme = engine["final_predictions"][budget]["scheme"]
        report_lines.extend(render_scheme_markdown(scheme, budget))
        report_lines.append("")
    report_lines.extend(
        [
            "## 规则文档核对",
            f"- 已从 DOCX 中提取到 {len(rule_text)} 个字符，用于核对投注与奖级规则。",
            "",
            "## 输出文件",
            f"- 预算摘要：{output_dir / 'budget_summary.csv'}",
            f"- 滚动回测明细：{output_dir / 'rolling_backtest.csv'}",
            f"- 高回报期数摘要：{output_dir / 'top_reward_issues.json'}",
            f"- 下一期详细方案：{output_dir / 'prediction_next_issue.json'}",
            f"- 下一期票号清单：{output_dir / 'prediction_tickets.csv'}",
            f"- 引擎状态：{output_dir / 'engine_state.json'}",
            f"- 自适应诊断：{output_dir / 'adaptive_diagnostics.json'}",
            f"- 官方增强摘要：{output_dir / 'official_enrichment.json'}",
            f"- 官方抓取缓存：{official_cache_path}",
        ]
    )
    (output_dir / "prediction_report.md").write_text("\n".join(report_lines), encoding="utf-8")


if __name__ == "__main__":
    main()
